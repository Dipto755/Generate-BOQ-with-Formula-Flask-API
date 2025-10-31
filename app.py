from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from pymongo import MongoClient
import traceback
import logging
from logging.handlers import RotatingFileHandler
import os
from datetime import datetime, timezone
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
import openpyxl
import uuid
import re
import redis
from functools import wraps
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock, Semaphore

load_dotenv()

# Configure logging
LOG_FOLDER = "logs"
SESSIONS_LOG_FOLDER = os.path.join(LOG_FOLDER, "sessions")
os.makedirs(LOG_FOLDER, exist_ok=True)
os.makedirs(SESSIONS_LOG_FOLDER, exist_ok=True)

def setup_session_logger():
    """Set up a new logger for each session with unique log file"""
    # Create session ID based on datetime
    session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Create logger
    logger = logging.getLogger(f'boq_calculator_{session_id}')
    logger.setLevel(logging.INFO)
    
    # Clear any existing handlers
    logger.handlers.clear()
    
    # Session-specific log file
    session_log_file = os.path.join(SESSIONS_LOG_FOLDER, f'session_{session_id}.log')
    file_handler = RotatingFileHandler(
        session_log_file,
        maxBytes=10*1024*1024,  # 10MB
        backupCount=3
    )
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    ))
    
    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s'
    ))
    
    # Add handlers
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger, session_id

# Initialize logger for the current session
logger, current_session_id = setup_session_logger()

app = Flask(__name__)
CORS(app)

# Add logger to Flask app
app.logger = logger

# MongoDB Configuration
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017/")
mongo_client = MongoClient(
    MONGO_URI,
    maxPoolSize=500,      # Changed from 100
    minPoolSize=50,       # Changed from 10
    maxIdleTimeMS=60000,
    waitQueueTimeoutMS=10000,
    serverSelectionTimeoutMS=10000,
    socketTimeoutMS=45000,
    connectTimeoutMS=10000
)
db = mongo_client[os.getenv("MONGO_DB_NAME")]

# Redis Configuration for caching
REDIS_HOST = os.getenv("REDIS_HOST", "localhost")
REDIS_PORT = int(os.getenv("REDIS_PORT", 6379))
# REDIS_DB = int(os.getenv("REDIS_DB", 0))
# REDIS_PASSWORD = os.getenv("REDIS_PASSWORD", None)

# Initialize Redis client
try:
    redis_client = redis.Redis(
        host=REDIS_HOST,
        port=REDIS_PORT,
        # db=REDIS_DB,
        # password=REDIS_PASSWORD,
        max_connections=100,
        decode_responses=True,  # Automatically decode responses to strings
        socket_connect_timeout=5
    )
    redis_client.ping()  # Test connection
    logger.info("Successfully connected to Redis")
except Exception as e:
    logger.warning(f"Redis connection failed: {e}. Caching will be disabled.")
    redis_client = None
    
# Limit concurrent LOOKUP operations
# lookup_semaphore = Semaphore(6)  # Max 3 LOOKUPs at once

# Collections for Main Carriageway formulas
main_carriageway_formulas_collection = db["formulas"]

# Collections for input file values (separate for each file type)
tcs_input_values_collection = db["tcs_input_values"]
pavement_input_values_collection = db["pavement_input_values"]
emb_height_values_collection = db["emb_height_values"]
tcs_schedule_values_collection = db["tcs_schedule_values"]

# Collections for different types of sessions and templates
app_sessions_collection = db["app_sessions"]  # For application runtime sessions
file_sessions_collection = db["file_sessions"]  # For file upload sessions
boq_templates_collection = db["boq_templates"]
calculated_main_carriageway_collection = db["calculated_main_carriageway_results"]  # For storing main carriageway calculations

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"
ALLOWED_EXTENSIONS = {"xlsx", "xls", "xlsm", "xlsb", "odf", "ods", "odt"}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# Global in-memory cache for input data during calculations
input_data_cache = {}
main_carriageway_formulas_cache = {}
cache_lock = Lock()

def load_input_data_to_memory(session_id):
    """Load all input data for a session into memory"""
    try:
        logger.info(f"Loading input data to memory for session: {session_id}")
        print(f"🔄 Loading input data to memory for session: {session_id}")
        
        # Collections to load
        collections_to_load = {
            "TCS Input.xlsx": tcs_input_values_collection,
            "Pavement Input.xlsx": pavement_input_values_collection, 
            "Emb Height.xlsx": emb_height_values_collection,
            "TCS Schedule.xlsx": tcs_schedule_values_collection
        }
        
        session_cache = {}
        total_loaded = 0
        
        for file_name, collection in collections_to_load.items():
            # Get all documents for this session and file
            cursor = collection.find({
                "session_id": session_id,
                "file_name": file_name
            })
            
            for doc in cursor:
                # Create key: {file_name}:{sheet}:{cell}
                cache_key = f"{file_name}:{doc['sheet']}:{doc['cell']}"
                session_cache[cache_key] = doc['value']
                total_loaded += 1
            
            logger.info(f"Loaded {total_loaded} cells for {file_name}")
            print(f"📦 Loaded {total_loaded} cells for {file_name}")
        
        # Store in global cache with session_id
        # with cache_lock:
        input_data_cache[session_id] = session_cache
        
        logger.info(f"Total input cells loaded to memory: {total_loaded}")
        print(f"✅ Total input cells loaded to memory: {total_loaded}")
        return True
        
    except Exception as e:
        logger.error(f"Error loading input data to memory: {e}")
        print(f"❌ Error loading input data to memory: {e}")
        return False

def get_input_data_from_memory(session_id, file_name, sheet_name, cell_address):
    """Get input data from in-memory cache"""
    try:
        # with cache_lock:
        if session_id not in input_data_cache:
            return None
        
        cache_key = f"{file_name}:{sheet_name}:{cell_address}"
        return input_data_cache[session_id].get(cache_key)
            
    except Exception as e:
        logger.error(f"Error getting data from memory cache: {e}")
        return None

def clear_input_data_from_memory(session_id):
    """Clear input data from memory cache for a session"""
    try:
        with cache_lock:
            if session_id in input_data_cache:
                del input_data_cache[session_id]
                logger.info(f"Cleared input data from memory for session: {session_id}")
                print(f"🗑️ Cleared input data from memory for session: {session_id}")
                return True
        return False
    except Exception as e:
        logger.error(f"Error clearing input data from memory: {e}")
        return False
    
def load_main_carriageway_formulas_to_memory(sheet_name=None):
    """Load Main Carriageway formulas to memory for faster access"""
    try:
        logger.info(f"Loading Main Carriageway formulas to memory for sheet: {sheet_name or 'ALL'}")
        print(f"🔄 Loading Main Carriageway formulas to memory...")
        
        # Build query - optionally filter by sheet
        query = {"file_name": "Main Carriageway.xlsx"}
        if sheet_name:
            query["sheet"] = sheet_name
        
        # Get all formula documents
        cursor = main_carriageway_formulas_collection.find(query)
        
        formulas_cache = {}
        total_loaded = 0
        
        for doc in cursor:
            sheet = doc.get('sheet')
            cell = doc.get('cell')
            
            if sheet and cell:
                # Create key: {sheet}:{cell}
                cache_key = f"{sheet}:{cell}"
                
                # Store the entire document (has formula, value, is_formula fields)
                formulas_cache[cache_key] = {
                    'is_formula': doc.get('is_formula', False),
                    'formula': doc.get('formula'),
                    'value': doc.get('value')
                }
                total_loaded += 1
        
        # Store in global cache
        if sheet_name:
            # Merge with existing cache if loading specific sheet
            main_carriageway_formulas_cache.update(formulas_cache)
        else:
            # Replace entire cache if loading all
            main_carriageway_formulas_cache.clear()
            main_carriageway_formulas_cache.update(formulas_cache)
        
        logger.info(f"Total Main Carriageway formulas loaded to memory: {total_loaded}")
        print(f"✅ Loaded {total_loaded} Main Carriageway formulas to memory")
        return True
        
    except Exception as e:
        logger.error(f"Error loading Main Carriageway formulas to memory: {e}")
        print(f"❌ Error loading formulas to memory: {e}")
        return False


def clear_main_carriageway_formulas_from_memory():
    """Clear Main Carriageway formulas from memory cache"""
    try:
        main_carriageway_formulas_cache.clear()
        logger.info("Cleared Main Carriageway formulas from memory")
        print("🗑️ Cleared Main Carriageway formulas from memory")
        return True
    except Exception as e:
        logger.error(f"Error clearing formulas from memory: {e}")
        return False


def get_collection_for_file(file_key):
    """Get the appropriate MongoDB collection for a file type"""
    collections = {
        "tcs_input": tcs_input_values_collection,
        "pavement_input": pavement_input_values_collection,
        "emb_height": emb_height_values_collection,
        "tcs_schedule": tcs_schedule_values_collection
    }
    return collections.get(file_key)


def parse_cell_address(cell_address):
    """Parse Excel cell address like 'AY5' into column and row"""
    match = re.match(r'([A-Z]+)(\d+)', cell_address.upper())
    if match:
        col_letter = match.group(1)
        row_num = int(match.group(2))
        return col_letter, row_num
    return None, None


def expand_range(range_str, current_sheet=None):
    """
    Expand Excel range like 'Q73:S73' to ['Q73', 'R73', 'S73']
    """
    try:
        if ':' not in range_str:
            return [range_str]
        
        # Remove sheet prefix if present: Sheet!Q73:S73
        if '!' in range_str:
            sheet_part, range_part = range_str.split('!', 1)
            prefix = sheet_part + '!'
        else:
            range_part = range_str
            prefix = ''
        
        # Parse start and end cells
        start_cell, end_cell = range_part.split(':')
        
        # Remove $ signs
        start_cell = start_cell.replace('$', '')
        end_cell = end_cell.replace('$', '')
        
        # Extract column and row
        start_col = ''.join(filter(str.isalpha, start_cell))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        end_col = ''.join(filter(str.isalpha, end_cell))
        end_row = int(''.join(filter(str.isdigit, end_cell)))
        
        # Convert column letters to numbers
        def col_to_num(col):
            num = 0
            for char in col:
                num = num * 26 + (ord(char.upper()) - ord('A') + 1)
            return num
        
        def num_to_col(num):
            col = ''
            while num > 0:
                num -= 1
                col = chr(num % 26 + ord('A')) + col
                num //= 26
            return col
        
        start_col_num = col_to_num(start_col)
        end_col_num = col_to_num(end_col)
        
        # Generate all cells in range
        cells = []
        for row in range(start_row, end_row + 1):
            for col_num in range(start_col_num, end_col_num + 1):
                col = num_to_col(col_num)
                cells.append(f"{prefix}{col}{row}")
        
        return cells
        
    except Exception as e:
        logger.error(f"Error expanding range '{range_str}': {e}")
        return [range_str]


def get_cell_value_from_db(session_id, file_name, sheet_name, cell_address, collection):
    """Retrieve cell value - try in-memory cache first, then fall back to MongoDB"""
    # First try to get from in-memory cache
    cached_value = get_input_data_from_memory(session_id, file_name, sheet_name, cell_address)
    
    if cached_value is not None:
        logger.debug(f"Memory cache HIT for input cell: {file_name}:{sheet_name}:{cell_address}")
        return cached_value
    
    # If not in memory cache, fall back to MongoDB
    try:
        cell_doc = collection.find_one({
            "session_id": session_id,
            "file_name": file_name,
            "sheet": sheet_name,
            "cell": cell_address
        })
        
        if cell_doc:
            value = cell_doc.get("value")
            logger.debug(f"MongoDB fallback for: {file_name}:{sheet_name}:{cell_address} = {value}")
            return value
        return 0
    except Exception as e:
        logger.error(f"Error retrieving cell value from MongoDB: {e}")
        return None


def col_letter_to_index(col):
    idx = 0
    for c in col:
        idx = idx * 26 + (ord(c) - ord('A') + 1)
    return idx


def index_to_col_letter(index):
    letters = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters = chr(rem + ord('A')) + letters
    return letters


def get_cache_key(session_id, cell_ref):
    """Generate Redis cache key for a cell calculation"""
    return f"calc:{session_id}:{cell_ref}"


def get_from_cache(session_id, cell_ref):
    """Get cached cell value from Redis"""
    if redis_client is None:
        return None
    
    try:
        cache_key = get_cache_key(session_id, cell_ref)
        cached_value = redis_client.get(cache_key)
        
        if cached_value is not None:
            logger.debug(f"Cache HIT for {cell_ref}")
            print(f"🎯 Cache HIT: {cell_ref}")
            # Try to convert to appropriate type
            try:
                # Try float first
                return float(cached_value)
            except ValueError:
                # Return as string if not a number
                return cached_value
        else:
            logger.debug(f"Cache MISS for {cell_ref}")
            return None
    except Exception as e:
        logger.error(f"Error getting from cache: {e}")
        return None


def set_to_cache(session_id, cell_ref, value, ttl=10800):
    """Set cell value to Redis cache with TTL (default 3 hour)"""
    if redis_client is None:
        return False
    
    try:
        cache_key = get_cache_key(session_id, cell_ref)
        redis_client.setex(cache_key, ttl, str(value))
        logger.debug(f"Cache SET for {cell_ref} = {value}")
        print(f"💾 Cache SET: {cell_ref} = {value}")
        return True
    except Exception as e:
        logger.error(f"Error setting to cache: {e}")
        return False


def clear_session_cache(session_id):
    """Clear all cached values for a session"""
    if redis_client is None:
        return 0
    
    try:
        pattern = f"calc:{session_id}:*"
        keys = redis_client.keys(pattern)
        if keys:
            count = redis_client.delete(*keys)
            logger.info(f"Cleared {count} cached values for session {session_id}")
            print(f"🗑️ Cleared {count} cached values")
            return count
        return 0
    except Exception as e:
        logger.error(f"Error clearing session cache: {e}")
        return 0
    

def flush_redis_cache():
    """Flush entire Redis database (removes all keys)"""
    if redis_client is None:
        return False, "Redis client not available"
    
    try:
        redis_client.flushdb()
        logger.info("Redis database flushed successfully")
        print("🗑️ Redis database flushed")
        return True, "Redis cache flushed successfully"
    except Exception as e:
        logger.error(f"Error flushing Redis cache: {e}")
        return False, str(e)


def generate_cells_in_range(start_cell, end_cell):
    """Generate cell coordinates from start to end (e.g., C3 -> C44 or A1 -> D4)"""
    start_col, start_row = parse_cell_address(start_cell)
    end_col, end_row = parse_cell_address(end_cell)
    if not start_col or not end_col:
        return []
    start_idx = col_letter_to_index(start_col)
    end_idx = col_letter_to_index(end_col)
    cells = []
    for col_idx in range(start_idx, end_idx + 1):
        col = index_to_col_letter(col_idx)
        for row in range(start_row, end_row + 1):
            cells.append(f"{col}{row}")
    return cells


def get_collection_and_filename_from_name(file_name):
    """Map a file name string to the collection and canonical file_name used in DB"""
    name = file_name.lower()
    if "pavement" in name:
        return pavement_input_values_collection, "Pavement Input.xlsx"
    if "tcs" in name and "schedule" not in name:
        return tcs_input_values_collection, "TCS Input.xlsx"
    if "emb" in name or "height" in name:
        return emb_height_values_collection, "Emb Height.xlsx"
    if "schedule" in name:
        return tcs_schedule_values_collection, "TCS Schedule.xlsx"
    return None, None


def get_range_values_from_db(session_id, file_name, sheet_name, start_cell, end_cell):
    """Retrieve a list of values for cells in a range from the appropriate collection"""
    collection, file_key = get_collection_and_filename_from_name(file_name)
    if collection is None:
        logger.debug(f"No collection mapping for file: {file_name}")
        return []
    cells = generate_cells_in_range(start_cell, end_cell)
    
    # Try memory cache first (BATCH READ - NO LOCK)
    session_cache = input_data_cache.get(session_id)
    if session_cache:
        values = []
        all_found = True
        
        for c in cells:
            cache_key = f"{file_key}:{sheet_name}:{c}"
            value = session_cache.get(cache_key)
            
            if value is not None:
                values.append(value)
            else:
                # Cache miss - need to fall back to MongoDB
                all_found = False
                break
        
        # If all cells found in cache, return immediately
        if all_found:
            logger.debug(f"Memory cache HIT for range {start_cell}:{end_cell} ({len(values)} cells)")
            return values
    
    # Fallback to MongoDB (BATCH QUERY - not individual queries)
    logger.debug(f"Memory cache MISS for range, querying MongoDB for {len(cells)} cells")
    
    query = {
        "session_id": session_id,
        "file_name": file_key,
        "sheet": sheet_name,
        "cell": {"$in": cells}  # Single query for ALL cells in range
    }
    
    cursor = collection.find(query)
    
    # Build lookup dict from MongoDB results
    cell_values = {doc['cell']: doc['value'] for doc in cursor}
    
    # Return values in correct order (matching cells list order)
    values = [cell_values.get(c, 0) for c in cells]
    
    logger.debug(f"Retrieved {len(values)} values from MongoDB")
    return values


def evaluate_lookup_function(formula, session_id, current_sheet=None):
    """
    Very small implementation for LOOKUP(lookup_value, lookup_range, result_range)
    Supports cross-file ranges and local cell refs for lookup_value (prefixed with current_sheet if needed).
    Performs exact-match lookup returning corresponding result_range value.
    """
    # with lookup_semaphore:
    try:
        # Extract content inside LOOKUP(...)
        match = re.match(r'LOOKUP\((.*)\)', formula, re.IGNORECASE)
        if not match:
            logger.debug("LOOKUP pattern not matched")
            return None
        content = match.group(1)
        parts = split_formula_parts(content)
        if len(parts) < 3:
            logger.debug("LOOKUP requires 3 arguments")
            return None

        lookup_value_str = parts[0].strip()
        lookup_range_str = parts[1].strip()
        result_range_str = parts[2].strip()
        
        print(f"LOOKUP args --> lookup value str: {lookup_value_str}, lookup range str: {lookup_range_str}, result range str: {result_range_str}")

        # Resolve lookup value (could be a cell reference or literal)
        # If it's a simple cell like $D7 or D7 and no sheet provided, prefix with current_sheet
        if re.match(r'^\$?[A-Z]{1,3}\$?\d+$', lookup_value_str) and '!' not in lookup_value_str:
            if not current_sheet:
                logger.debug("No current_sheet provided for relative cell ref in LOOKUP")
                return None
            lookup_value = resolve_cell_reference(f"{current_sheet}!{lookup_value_str}", session_id)
            print(f"``````LOOKUP relative cell ref resolved to: {lookup_value}")
        else:
            # Use existing resolver to handle quoted strings or cross-file refs
            lookup_value = (
                resolve_value(lookup_value_str, session_id)
                if '!' not in lookup_value_str
                else resolve_cell_reference(lookup_value_str, session_id)
            )

        logger.debug(f"LOOKUP value resolved to: {lookup_value}")
        print("_________________________ Lookup value:", lookup_value)

        # Helper to parse a range like '[TCS Input.xlsx]Input!$C$3:$C$44' or 'Input!C3:C44'
        def parse_range_str(range_str):
            # Cross-file?
            if range_str.startswith("'[") or range_str.startswith("["):
                m = re.match(r"'?\[([^\]]+)\]'?(.+)", range_str)
                if not m:
                    return None
                file_name = m.group(1)
                rest = m.group(2)
                if '!' not in rest:
                    return None
                sheet_part, cells_part = rest.split('!', 1)
                sheet_name = sheet_part.strip("'")
            else:
                # local sheet included?
                if '!' in range_str:
                    sheet_part, cells_part = range_str.split('!', 1)
                    sheet_name = sheet_part.strip("'")
                    file_name = None
                else:
                    # No sheet/file — not supported for ranges
                    return None
            cells_part = cells_part.replace('$', '')
            if ':' not in cells_part:
                return None
            start_cell, end_cell = cells_part.split(':', 1)
            return file_name, sheet_name, start_cell.strip(), end_cell.strip()

        lookup_parsed = parse_range_str(lookup_range_str)
        result_parsed = parse_range_str(result_range_str)
        if not lookup_parsed or not result_parsed:
            logger.debug("Failed to parse LOOKUP ranges")
            return None

        lookup_file, lookup_sheet, lookup_start, lookup_end = lookup_parsed
        result_file, result_sheet, result_start, result_end = result_parsed

        # If ranges come from files, use DB; otherwise fail
        if lookup_file:
            lookup_values = get_range_values_from_db(session_id, lookup_file, lookup_sheet, lookup_start, lookup_end)
        else:
            logger.debug("Local range lookups not implemented")
            lookup_values = []

        if result_file:
            result_values = get_range_values_from_db(session_id, result_file, result_sheet, result_start, result_end)
        else:
            logger.debug("Local result ranges not implemented")
            result_values = []

        # Perform exact match lookup
        for idx, lv in enumerate(lookup_values):
            if lv == lookup_value:
                # Return corresponding result value if exists
                if idx < len(result_values):
                    logger.debug(f"LOOKUP matched at index {idx}: returning {result_values[idx]}")
                    return result_values[idx]
                break

        logger.debug("LOOKUP did not find a match")
        return None
    except Exception as e:
        logger.error(f"Error evaluating LOOKUP: {str(e)}", exc_info=True)
        return None


def evaluate_excel_formula(formula, session_id, current_sheet=None):
    """
    Evaluate Excel formula by resolving cell references and functions
    
    Supports:
    - Direct cell references: =Quantity!$GY$5091
    - Cross-file references: ='[Pavement input.xlsx]'Summary!$E$9
    - IF functions: =IF(condition, true_value, false_value)
    - LOOKUP: =LOOKUP(lookup_value, lookup_range, result_range)
    - Basic arithmetic: +, -, *, /
    """
    try:
        # Create cache key for this formula
        formula_cache_key = f"formula:{current_sheet}:{formula}"
        
        # Check cache first
        cached_result = get_from_cache(session_id, formula_cache_key)
        if cached_result is not None:
            logger.info(f"Formula cache HIT: {formula}")
            return cached_result
        
        logger.info(f"Starting formula evaluation. Formula: {formula}, Session ID: {session_id}, Sheet: {current_sheet}")
        print(f"📝 Processing formula: {formula}")

        # Remove leading '=' if present
        formula = formula.strip()
        if formula.startswith('='):
            formula = formula[1:]
            logger.debug("Removed leading '=' from formula")
        
        # Handle IF function
        # But first check if there's arithmetic after the IF (like IF(...)/1000)
        if formula.upper().startswith('IF('):
            # Check if there are arithmetic operations outside the IF function
            # Find the closing parenthesis of the IF function
            try:
                paren_count = 0
                if_end_index = -1
                for i, char in enumerate(formula):
                    if char == '(':
                        paren_count += 1
                    elif char == ')':
                        paren_count -= 1
                        if paren_count == 0:
                            if_end_index = i
                            break
                
                # Check if there's anything after the IF function closes
                if if_end_index > 0 and if_end_index < len(formula) - 1:
                    remaining = formula[if_end_index + 1:].strip()
                    if remaining and re.search(r'^[+\-*/]', remaining):
                        # There's arithmetic after the IF: IF(...)/1000
                        print(f"🔄 IF function has arithmetic after it: {remaining}")
                        logger.info("IF function has arithmetic operations after it, evaluating IF first")
                        
                        # Extract just the IF part
                        if_part = formula[:if_end_index + 1]
                        
                        # Evaluate IF function first
                        if_result = evaluate_if_function(if_part, session_id, current_sheet=current_sheet)
                        
                        if if_result is not None:
                            # Now do the arithmetic on the result
                            try:
                                # First, resolve any Excel functions in the remaining part
                                remaining_resolved = remaining
                                
                                # Handle SUM functions BEFORE cell resolution
                                if 'SUM(' in remaining_resolved.upper():
                                    sum_pattern = r'SUM\([^)]+\)'
                                    def replace_sum(match):
                                        sum_result = evaluate_sum_function(match.group(0), session_id, current_sheet=current_sheet)
                                        return str(sum_result) if sum_result is not None else match.group(0)
                                    remaining_resolved = re.sub(sum_pattern, replace_sum, remaining_resolved, flags=re.IGNORECASE)
                                
                                # Handle AVERAGE functions BEFORE cell resolution
                                if 'AVERAGE(' in remaining_resolved.upper():
                                    avg_pattern = r'AVERAGE\([^)]+\)'
                                    def replace_avg(match):
                                        avg_result = evaluate_average_function(match.group(0), session_id, current_sheet=current_sheet)
                                        return str(avg_result) if avg_result is not None else match.group(0)
                                    remaining_resolved = re.sub(avg_pattern, replace_avg, remaining_resolved, flags=re.IGNORECASE)
                                
                                # Resolve any cell references in the remaining part
                                remaining_resolved = resolve_all_cell_references(remaining_resolved, session_id, current_sheet=current_sheet)
                                print(f"🔍 Remaining after cell resolution: {remaining_resolved}")
                                
                                # Handle nested IF functions AFTER cell resolution
                                # Use a more robust approach that handles deeply nested parentheses
                                while 'IF(' in remaining_resolved.upper():
                                    # Find the start of the IF function
                                    if_start = remaining_resolved.upper().find('IF(')
                                    if if_start == -1:
                                        break
                                    
                                    # Find the matching closing parenthesis by counting
                                    paren_count = 1
                                    if_end = -1
                                    for i in range(if_start + 3, len(remaining_resolved)):  # Start after "IF"
                                        if remaining_resolved[i] == '(':
                                            paren_count += 1
                                        elif remaining_resolved[i] == ')':
                                            paren_count -= 1
                                            if paren_count == 0:
                                                if_end = i
                                                break

                                    
                                    if if_end == -1:
                                        logger.error("Could not find matching parenthesis for IF function")
                                        break
                                    
                                    # Extract the IF function
                                    if_formula = remaining_resolved[if_start:if_end + 1]
                                    print(f"  🔢 Found nested IF function: {if_formula}")
                                    
                                    # Evaluate it
                                    if_result = evaluate_if_function(if_formula, session_id, current_sheet)
                                    
                                    if if_result is not None:
                                        # Replace the IF function with its result
                                        remaining_resolved = remaining_resolved[:if_start] + str(if_result) + remaining_resolved[if_end + 1:]
                                        print(f"🔍 After IF replacement: {remaining_resolved}")
                                    else:
                                        logger.warning(f"IF function evaluation returned None: {if_formula}")
                                        break
                                
                                # Now evaluate the complete arithmetic expression
                                result = eval(f"{if_result}{remaining_resolved}", {"__builtins__": {}}, {})
                                logger.info(f"IF with arithmetic evaluated. Result: {result}")
                                print(f"✅ IF with arithmetic result: {result}")
                                return result
                            except Exception as e:
                                logger.error(f"Error in arithmetic after IF: {e}")
                                return None
                        else:
                            logger.warning("IF function returned None, cannot do arithmetic")
                            return None
                    else:
                        # Pure IF function without trailing arithmetic
                        logger.info("Detected IF function, delegating to evaluate_if_function")
                        print("🔄 Processing IF function...")
                        result = evaluate_if_function(formula, session_id, current_sheet=current_sheet)
                        if result is not None:
                            logger.info(f"IF function evaluated successfully. Result: {result}")
                            print(f"✅ IF function result: {result}")
                        else:
                            logger.warning("IF function evaluation returned None")
                            print("⚠️ IF function evaluation failed")
                        return result
                else:
                    # Pure IF function
                    logger.info("Detected IF function, delegating to evaluate_if_function")
                    print("🔄 Processing IF function...")
                    result = evaluate_if_function(formula, session_id, current_sheet=current_sheet)
                    if result is not None:
                        logger.info(f"IF function evaluated successfully. Result: {result}")
                        print(f"✅ IF function result: {result}")
                    else:
                        logger.warning("IF function evaluation returned None")
                        print("⚠️ IF function evaluation failed")
                    return result
            except Exception as e:
                logger.error(f"Error parsing IF function: {e}")
                # Fall through to arithmetic handling

        # Handle LOOKUP function
        if formula.upper().startswith('LOOKUP('):
            logger.info("Detected LOOKUP function, delegating to evaluate_lookup_function")
            print("🔄 Processing LOOKUP function...")
            result = evaluate_lookup_function(formula, session_id, current_sheet=current_sheet)
            if result is not None:
                logger.info(f"LOOKUP function result: {result}")
                print(f"✅ LOOKUP result: {result}")
            else:
                logger.warning("LOOKUP evaluation returned None")
                print("⚠️ LOOKUP evaluation failed")
            return result
        
        # Handle ROUNDUP function
        if formula.upper().startswith('ROUNDUP('):
            logger.info("Detected ROUNDUP function")
            print("🔄 Processing ROUNDUP function...")
            result = evaluate_roundup_function(formula, session_id, current_sheet=current_sheet)
            if result is not None:
                logger.info(f"ROUNDUP function result: {result}")
                print(f"✅ ROUNDUP result: {result}")
            return result
        
        # Handle ROUND function
        if formula.upper().startswith('ROUND('):
            logger.info("Detected ROUND function")
            print("🔄 Processing ROUND function...")
            result = evaluate_round_function(formula, session_id, current_sheet=current_sheet)
            if result is not None:
                logger.info(f"ROUND function result: {result}")
                print(f"✅ ROUND result: {result}")
            return result

        # Handle SQRT function
        if formula.upper().startswith('SQRT('):
            logger.info("Detected SQRT function")
            print("🔄 Processing SQRT function...")
            result = evaluate_sqrt_function(formula, session_id, current_sheet=current_sheet)
            if result is not None:
                logger.info(f"SQRT function result: {result}")
                print(f"✅ SQRT result: {result}")
            return result
        
        # Handle IFERROR function
        if formula.upper().startswith('IFERROR('):
            logger.info("Detected IFERROR function")
            print("🔄 Processing IFERROR function...")
            result = evaluate_iferror_function(formula, session_id, current_sheet=current_sheet)
            if result is not None:
                logger.info(f"IFERROR function result: {result}")
                print(f"✅ IFERROR result: {result}")
            else:
                logger.warning("IFERROR evaluation returned None")
                print("⚠️ IFERROR evaluation failed")
            return result
        
        
        # Handle complex expressions with SUM/AVERAGE and arithmetic
        # Replace SUM(...) and AVERAGE(...) calls with their evaluated results before arithmetic
        temp_formula = formula
        sum_pattern = r'SUM\([^)]+\)'
        avg_pattern = r'AVERAGE\([^)]+\)'

        def replace_sum(match):
            sum_result = evaluate_sum_function(match.group(0), session_id, current_sheet=current_sheet)
            return str(sum_result) if sum_result is not None else match.group(0)

        def replace_avg(match):
            avg_result = evaluate_average_function(match.group(0), session_id, current_sheet=current_sheet)
            return str(avg_result) if avg_result is not None else match.group(0)

        # Replace SUM functions
        import re as regex_module
        temp_formula = regex_module.sub(sum_pattern, replace_sum, temp_formula, flags=regex_module.IGNORECASE)

        # Replace AVERAGE functions
        temp_formula = regex_module.sub(avg_pattern, replace_avg, temp_formula, flags=regex_module.IGNORECASE)

        # If formula was modified, use the modified version for further processing
        if temp_formula != formula:
            formula = temp_formula
            logger.info(f"Formula after SUM/AVERAGE evaluation: {formula}")
            print(f"🔄 Formula after function evaluation: {formula}")
        
        # Handle direct cell reference (e.g., Quantity!$GY$5091)
        if '!' in formula and not any(op in formula for op in ['+', '-', '*', '/', '(', ')']):
            logger.info(f"Detected direct cell reference: {formula}")
            print(f"🔄 Resolving cell reference: {formula}")
            result = resolve_cell_reference(formula, session_id)
            if result is not None:
                logger.info(f"Cell reference resolved. Value: {result}")
                print(f"✅ Cell value retrieved: {result}")
            else:
                logger.warning(f"Failed to resolve cell reference: {formula}")
                print("⚠️ Cell reference resolution failed")
            return result
        
        # Handle arithmetic expressions
        # Replace cell references with their values
        logger.info("Processing arithmetic expression")
        print("🔄 Processing arithmetic expression...")
        
        # Resolve all cell references in the formula
        # Pass current_sheet so plain cell refs like $D7 get prefixed correctly
        # But cross-file/sheet refs like '[File]Sheet'!A1 won't be modified
        resolved_formula = resolve_all_cell_references(formula, session_id, current_sheet=current_sheet)
        logger.debug(f"Resolved formula after replacing cell references: {resolved_formula}")
        print(f"🔍 Final resolved formula: {resolved_formula}")
        
        # Evaluate the expression safely
        try:
            result = safe_eval(resolved_formula)
            if result is not None:
                logger.info(f"Formula evaluated successfully. Result: {result}")
                print(f"✅ Formula result: {result}")
            else:
                logger.warning("Formula evaluation returned None")
                print("⚠️ Formula evaluation failed")
            return result
        except Exception as e:
            logger.error(f"Error in safe_eval: {str(e)}", exc_info=True)
            print(f"❌ Error evaluating expression: {str(e)}")
            return None
            
    except Exception as e:
        logger.error(f"Error evaluating formula '{formula}': {str(e)}", exc_info=True)
        print(f"❌ Formula evaluation failed: {str(e)}")
        return None


def evaluate_if_function(formula, session_id, current_sheet=None):
    """Evaluate Excel IF function: IF(condition, true_value, false_value)

    Accepts optional current_sheet so relative references inside the IF (including OR/AND)
    are resolved against the correct sheet.
    """
    try:
        # Extract the content inside IF()
        match = re.match(r'IF\((.*)\)', formula, re.IGNORECASE)
        if not match:
            return None
        
        content = match.group(1)
        print("***************************** IF function content:", content)
        
        # Split by commas (careful with nested functions)
        parts = split_formula_parts(content)
        
        if len(parts) != 3:
            return None
        
        condition_str, true_value_str, false_value_str = parts
        
        print("---------------- IF parts ----------------")
        print(f"Condition str --> {condition_str}")
        print(f"True value str --> {true_value_str}")
        print(f"False value str --> {false_value_str}")
        
        # Resolve condition - but don't resolve cells inside functions
        condition_str = condition_str.strip()
        # Only resolve if not containing Excel functions that will handle their own cell resolution
        if not any(func in condition_str.upper() for func in ['SUM(', 'AVERAGE(', 'LOOKUP(', 'IF(', 'OR(', 'AND(']):
            condition_str = resolve_all_cell_references(condition_str, session_id, current_sheet=current_sheet)

        # Check if condition contains OR function
        if condition_str.upper().startswith('OR('):
            condition_result = evaluate_or_function(condition_str, session_id)
        else:
            # Evaluate condition
            condition_result = evaluate_condition(condition_str)
        
        # Choose branch
        branch_raw = true_value_str.strip() if condition_result else false_value_str.strip()
        
        # Remove surrounding parentheses if present
        if branch_raw.startswith('(') and branch_raw.endswith(')'):
            branch_raw = branch_raw[1:-1].strip()
        
        # Check if branch contains ANY Excel functions (including SUM and AVERAGE)
        # These functions need to be evaluated through evaluate_excel_formula
        if re.search(r'\b(IF|OR|AND|LOOKUP|SUM|AVERAGE|ROUNDUP|SQRT|IFERROR)\s*\(', branch_raw, re.IGNORECASE):
            print("🔄 Branch contains Excel functions, evaluating through evaluate_excel_formula")
            return evaluate_excel_formula(branch_raw, session_id, current_sheet=current_sheet)
        
        # If branch is just a nested IF without arithmetic
        if branch_raw.upper().startswith('IF('):
            print("🔄 Branch is nested IF, evaluating recursively")
            return evaluate_if_function(branch_raw, session_id, current_sheet=current_sheet)
        
        # No Excel functions detected - safe to resolve cell references
        branch_expr = resolve_all_cell_references(branch_raw, session_id, current_sheet=current_sheet).strip()
        
        # If branch contains arithmetic or parentheses, try safe_eval
        if re.search(r'[+\-*/()]', branch_expr):
            try:
                val = safe_eval(branch_expr)
                if val is not None:
                    return val
            except Exception:
                logger.debug(f"safe_eval failed for branch expression: {branch_expr}")
            # fallback to resolve_value of original branch (in case of strings)
            return resolve_value(branch_expr, session_id, current_sheet)
        
        # No arithmetic — resolve as value (literal, cell ref resolved already)
        return resolve_value(branch_expr, session_id, current_sheet)
            
    except Exception as e:
        print(f"Error evaluating IF function: {e}")
        logger.error(f"Error evaluating IF function: {e}", exc_info=True)
        return None
    
def evaluate_or_function(formula, session_id):
    """Evaluate Excel OR function: OR(condition1, condition2, ...)"""
    try:
        match = re.match(r'OR\((.*)\)', formula, re.IGNORECASE)
        if not match:
            return None
        
        content = match.group(1)
        parts = split_formula_parts(content)
        
        # Evaluate each condition
        for part in parts:
            condition_str = resolve_all_cell_references(part.strip(), session_id)
            if evaluate_condition(condition_str):
                return True
        
        return False
    except Exception as e:
        logger.error(f"Error evaluating OR function: {e}")
        return False
    

def evaluate_sum_function(formula, session_id, current_sheet=None):
    """Evaluate Excel SUM function: SUM(range) or SUM(val1, val2, ...)"""
    try:
        match = re.match(r'SUM\(((?:[^()]+|\([^()]*\))*)\)', formula, re.IGNORECASE)
        if not match:
            return None

        content = match.group(1)  # Now group(1) exists!
        
        print("############################################ SUM function content:", content)
        
        # Check if it's a range or individual values
        parts = split_formula_parts(content)
        
        total = 0
        for part in parts:
            part = part.strip()
            
            # Check if it's a range
            if ':' in part:
                # Expand range
                cells = expand_range(part, current_sheet)
                
                # Sum all cells in range
                for cell in cells:
                    # Resolve cell reference
                    value = resolve_cell_reference(cell, session_id, current_sheet)
                    if value is not None:
                        try:
                            total += float(value)
                        except (ValueError, TypeError):
                            pass
            else:
                # Single cell or value
                resolved = resolve_all_cell_references(part, session_id, current_sheet)
                try:
                    total += float(resolved)
                except (ValueError, TypeError):
                    pass
        
        return total
        
    except Exception as e:
        logger.error(f"Error evaluating SUM function: {e}")
        return None
    
    
def evaluate_average_function(formula, session_id, current_sheet=None):
    """Evaluate Excel AVERAGE function: AVERAGE(range) or AVERAGE(val1, val2, ...)"""
    try:
        match = re.match(r'AVERAGE\((.*)\)', formula, re.IGNORECASE)
        if not match:
            return None
        
        content = match.group(1)
        parts = split_formula_parts(content)
        
        values = []
        for part in parts:
            part = part.strip()
            
            # Check if it's a range
            if ':' in part:
                cells = expand_range(part, current_sheet)
                
                for cell in cells:
                    value = resolve_cell_reference(cell, session_id, current_sheet)
                    if value is not None:
                        try:
                            values.append(float(value))
                        except (ValueError, TypeError):
                            pass
            else:
                # Single cell or value
                resolved = resolve_all_cell_references(part, session_id, current_sheet)
                try:
                    values.append(float(resolved))
                except (ValueError, TypeError):
                    pass
        
        if len(values) == 0:
            return 0
        
        return sum(values) / len(values)
        
    except Exception as e:
        logger.error(f"Error evaluating AVERAGE function: {e}")
        return None


def evaluate_roundup_function(formula, session_id, current_sheet=None):
    """Evaluate Excel ROUNDUP function: ROUNDUP(number, num_digits)"""
    try:
        match = re.match(r'ROUNDUP\((.*)\)', formula, re.IGNORECASE)
        if not match:
            return None
        
        content = match.group(1)
        parts = split_formula_parts(content)
        
        print(f"^^^^^^^^^^^^^^^^^^^^^^^^^^^^ ROUNDUP function got {len(parts)}  <<<<----- parts:", parts)
        
        if len(parts) != 2:
            return None
        
        number_str, digits_str = parts
        
        # Resolve and evaluate the number part (might contain nested formulas)
        number_resolved = resolve_all_cell_references(number_str.strip(), session_id, current_sheet)
        
        # Check if it contains nested functions
        if any(func in number_resolved.upper() for func in ['IF(', 'SQRT(', 'SUM(', 'AVERAGE(']):
            number = evaluate_excel_formula(number_resolved, session_id, current_sheet)
        else:
            number = safe_eval(number_resolved)
        
        # Resolve digits
        digits_resolved = resolve_all_cell_references(digits_str.strip(), session_id, current_sheet)
        digits = safe_eval(digits_resolved)
        
        if number is None or digits is None:
            return None
        
        import math
        multiplier = 10 ** int(digits)
        return math.ceil(float(number) * multiplier) / multiplier
        
    except Exception as e:
        logger.error(f"Error evaluating ROUNDUP function: {e}")
        return None


def evaluate_sqrt_function(formula, session_id, current_sheet=None):
    """Evaluate Excel SQRT function: SQRT(number)"""
    try:
        match = re.match(r'SQRT\((.*)\)', formula, re.IGNORECASE)
        if not match:
            return None
        
        content = match.group(1)
        
        # Resolve cell references
        resolved = resolve_all_cell_references(content.strip(), session_id, current_sheet)
        
        # Evaluate the expression
        number = safe_eval(resolved)
        
        if number is None:
            return None
        
        import math
        return math.sqrt(float(number))
        
    except Exception as e:
        logger.error(f"Error evaluating SQRT function: {e}")
        return None
    

def evaluate_round_function(formula, session_id, current_sheet=None):
    """Evaluate Excel ROUND function: ROUND(number, num_digits)"""
    try:
        match = re.match(r'ROUND\((.*)\)', formula, re.IGNORECASE)
        if not match:
            return None
        
        content = match.group(1)
        parts = split_formula_parts(content)
        
        print(f"^^^^^^^^^^^^^^^^^^^^^^^^^^^^ ROUND function got {len(parts)}  <<<<----- parts:", parts)
        
        if len(parts) != 2:
            return None
        
        number_str, digits_str = parts
        
        # Resolve and evaluate the number part (might contain nested formulas)
        number_resolved = resolve_all_cell_references(number_str.strip(), session_id, current_sheet)
        
        # Check if it contains nested functions
        if any(func in number_resolved.upper() for func in ['IF(', 'SQRT(', 'SUM(', 'AVERAGE(']):
            number = evaluate_excel_formula(number_resolved, session_id, current_sheet)
        else:
            number = safe_eval(number_resolved)
        
        # Resolve digits
        digits_resolved = resolve_all_cell_references(digits_str.strip(), session_id, current_sheet)
        digits = safe_eval(digits_resolved)
        
        if number is None or digits is None:
            return None
        
        # Use Python's built-in round function
        return round(float(number), int(digits))
        
    except Exception as e:
        logger.error(f"Error evaluating ROUND function: {e}")
        return None
    
    
def evaluate_iferror_function(formula, session_id, current_sheet=None):
    """Evaluate Excel IFERROR function: IFERROR(value, value_if_error)"""
    try:
        match = re.match(r'IFERROR\((.*)\)', formula, re.IGNORECASE)
        if not match:
            return None
        
        content = match.group(1)
        print("############################################ IFERROR function content:", content)
        
        # Split into value and value_if_error parts
        parts = split_formula_parts(content)
        
        if len(parts) != 2:
            logger.error(f"IFERROR requires exactly 2 arguments, got {len(parts)}")
            return None
        
        value_str, value_if_error_str = parts
        
        print("---------------- IFERROR parts ----------------")
        print(f"Value to try: {value_str}")
        print(f"Value if error: {value_if_error_str}")
        
        # Try to evaluate the first argument
        try:
            result = evaluate_excel_formula(value_str.strip(), session_id, current_sheet=current_sheet)
            
            if result is not None:
                logger.info(f"IFERROR: Value evaluated successfully = {result}")
                print(f"✅ IFERROR: Value succeeded = {result}")
                return result
            else:
                # If result is None, treat as error and use fallback
                logger.info("IFERROR: Value returned None, using fallback")
                print("⚠️ IFERROR: Value returned None, using fallback")
                fallback = evaluate_excel_formula(value_if_error_str.strip(), session_id, current_sheet=current_sheet)
                return fallback
                
        except Exception as e:
            # Error occurred, use value_if_error
            logger.info(f"IFERROR: Error occurred ({str(e)}), using fallback value")
            print("⚠️ IFERROR: Error occurred, using fallback")
            
            fallback = evaluate_excel_formula(value_if_error_str.strip(), session_id, current_sheet=current_sheet)
            return fallback
            
    except Exception as e:
        logger.error(f"Error evaluating IFERROR function: {e}")
        print(f"❌ IFERROR evaluation failed: {e}")
        return None


def split_formula_parts(content):
    """Split formula by commas, respecting nested parentheses and quotes"""
    parts = []
    current_part = ""
    paren_depth = 0
    in_quotes = False
    quote_char = None
    
    for char in content:
        if char in ['"', "'"]:
            if not in_quotes:
                in_quotes = True
                quote_char = char
            elif char == quote_char:
                in_quotes = False
                quote_char = None
            current_part += char
        elif char == '(' and not in_quotes:
            paren_depth += 1
            current_part += char
        elif char == ')' and not in_quotes:
            paren_depth -= 1
            current_part += char
        elif char == ',' and paren_depth == 0 and not in_quotes:
            parts.append(current_part)
            current_part = ""
        else:
            current_part += char
    
    if current_part:
        parts.append(current_part)
    
    return parts


def evaluate_condition(condition_str):
    """Evaluate a condition string (supports OR(...), AND(...), basic comparisons).

    Assumes any cell references have been resolved beforehand (strings or numbers).
    Supports nested OR/AND calls and simple comparisons for strings and numbers.
    """
    try:
        s = condition_str.strip()

        # Support OR(...) and AND(...)
        upper = s.upper()
        if upper.startswith('OR(') and s.endswith(')'):
            inner = s[s.find('(') + 1:-1]
            parts = split_formula_parts(inner)
            for part in parts:
                if evaluate_condition(part.strip()):
                    return True
            return False

        if upper.startswith('AND(') and s.endswith(')'):
            inner = s[s.find('(') + 1:-1]
            parts = split_formula_parts(inner)
            for part in parts:
                if not evaluate_condition(part.strip()):
                    return False
            return True

        # Unwrap surrounding parentheses
        if s.startswith('(') and s.endswith(')'):
            return evaluate_condition(s[1:-1].strip())

        # Handle string equality using single '=' as in Excel
        if '=' in s and '==' not in s:
            parts = s.split('=', 1)
            if len(parts) == 2:
                left = parts[0].strip().strip('"').strip("'")
                right = parts[1].strip().strip('"').strip("'")
                return left == right

        # Handle numeric comparisons in order to avoid prefix issues
        for op in ['>=', '<=', '!=', '>', '<']:
            if op in s:
                parts = s.split(op, 1)
                if len(parts) == 2:
                    left_str = parts[0].strip().strip('"').strip("'")
                    right_str = parts[1].strip().strip('"').strip("'")
                    try:
                        left = float(left_str)
                        right = float(right_str)
                        if op == '>=':
                            return left >= right
                        if op == '<=':
                            return left <= right
                        if op == '>':
                            return left > right
                        if op == '<':
                            return left < right
                        if op == '!=':
                            return left != right
                    except Exception:
                        logger.debug(f"Could not convert to float in condition: '{left_str}' or '{right_str}'")
                        return False

        # Boolean literals
        if s.upper() in ('TRUE', 'FALSE'):
            return s.upper() == 'TRUE'

        # Numeric truthiness (non-zero is True)
        try:
            num = float(s)
            return num != 0
        except Exception:
            logger.debug(f"Unable to evaluate condition: '{s}'")
            return False

    except Exception as e:
        logger.error(f"Error evaluating condition: {e}", exc_info=True)
        return False


def resolve_value(value_str, session_id, current_sheet=None):
    """Resolve a value (could be string literal, cell reference, or number)"""
    value_str = value_str.strip()
    
    # String literal
    if (value_str.startswith('"') and value_str.endswith('"')) or \
       (value_str.startswith("'") and value_str.endswith("'")):
        return value_str[1:-1]
    
    # Cell reference
    if '!' in value_str:
        return resolve_cell_reference(value_str, session_id, current_sheet)
    
    # Number
    try:
        return float(value_str)
    except ValueError as e:
        logger.debug(f"Value {value_str} could not be converted to float: {str(e)}")
        return value_str


def resolve_cell_reference(cell_ref, session_id, current_sheet=None):
    """
    Resolve a cell reference to its value
    Formats supported:
    - Quantity!$GY$5091 (same file, different sheet)
    - '[Pavement input.xlsx]'Summary!$E$9 (cross-file reference)
    """
    
    # If it's a plain cell reference (no sheet specified) and current_sheet is provided
    if '!' not in cell_ref and current_sheet:
        cell_ref = f"{current_sheet}!{cell_ref}"
    
    # Check cache first
    cached_value = get_from_cache(session_id, cell_ref)
    if cached_value is not None:
        return cached_value
    
    try:
        cell_ref = cell_ref.strip()
        
        # Check if it's a cross-file reference
        if cell_ref.startswith("'[") or cell_ref.startswith("["):
            # Extract file name
            match = re.match(r"'?\[([^\]]+)\]'?(.+)", cell_ref)
            if match:
                file_name = match.group(1)
                rest = match.group(2)
                print("---------------------------", file_name, rest)
                
                # Parse sheet and cell
                if '!' in rest:
                    sheet_and_cell = rest.split('!', 1)
                    sheet_name = sheet_and_cell[0].strip("'")
                    cell_address = sheet_and_cell[1].strip().replace('$', '')
                    
                    # Map file name to collection
                    collection = None
                    file_key = None
                    if "pavement" in file_name.lower():
                        collection = pavement_input_values_collection
                        file_key = "Pavement Input.xlsx"
                    elif "tcs" in file_name.lower() and "schedule" not in file_name.lower():
                        collection = tcs_input_values_collection
                        file_key = "TCS Input.xlsx"
                    elif "emb" in file_name.lower() or "height" in file_name.lower():
                        collection = emb_height_values_collection
                        file_key = "Emb Height.xlsx"
                    elif "schedule" in file_name.lower():
                        collection = tcs_schedule_values_collection
                        file_key = "TCS Schedule.xlsx"
                    
                    print("=====================", collection, file_key)
                    if collection is not None:
                        value = get_cell_value_from_db(session_id, file_key, sheet_name, cell_address, collection)
                        # Cache the value before returning
                        if value is not None:
                            set_to_cache(session_id, cell_ref, value)
                        return value
        else:
            # Same-file reference (Main Carriageway)
            if '!' in cell_ref:
                parts = cell_ref.split('!', 1)
                sheet_name = parts[0].strip("'")
                cell_address = parts[1].strip().replace('$', '')
                
                # Try memory cache first
                cache_key = f"{sheet_name}:{cell_address}"
                formula_doc = main_carriageway_formulas_cache.get(cache_key)
                
                if formula_doc:
                    logger.debug(f"Memory cache HIT for Main Carriageway: {sheet_name}!{cell_address}")
                    print(f"🎯 Formula cache HIT: {sheet_name}!{cell_address}")
                else:
                    # Fallback to MongoDB
                    logger.debug(f"Memory cache MISS, querying MongoDB for: {sheet_name}!{cell_address}")
                    print(f"📊 Formula cache MISS, querying MongoDB...")
                    
                    formula_doc_mongo = main_carriageway_formulas_collection.find_one({
                        "file_name": "Main Carriageway.xlsx",
                        "sheet": sheet_name,
                        "cell": cell_address
                    })
                    
                    if formula_doc_mongo:
                        # Convert to cache format
                        formula_doc = {
                            'is_formula': formula_doc_mongo.get('is_formula', False),
                            'formula': formula_doc_mongo.get('formula'),
                            'value': formula_doc_mongo.get('value')
                        }

                if formula_doc:
                    # Check if it's a formula or a value
                    if formula_doc.get("is_formula"):
                        # It's a formula - evaluate it recursively
                        formula = formula_doc.get("formula")
                        if formula:
                            logger.debug(f"Cell {cell_address} contains formula: {formula}")
                            print(f"Formula doc found for {current_sheet}!{cell_address}: {formula_doc}")
                            result = evaluate_excel_formula(formula, session_id, current_sheet=sheet_name)
                            # Cache the calculated value before returning
                            if result is not None:
                                set_to_cache(session_id, cell_ref, result)
                            return result
                    else:
                        # It's a value - return directly
                        value = formula_doc.get("value")
                        logger.debug(f"Cell {cell_address} contains value: {value}")
                        print(f"Value doc found for {current_sheet}!{cell_address}: value={value}")
                        # Cache the value before returning
                        if value is not None:
                            set_to_cache(session_id, cell_ref, value)
                        return value
                
        return None
    except Exception as e:
        print(f"Error resolving cell reference '{cell_ref}': {e}")
        return None


def resolve_all_cell_references(formula, session_id, current_sheet=None):
    """Replace all cell references in a formula with their values.

    Supports:
      - '[File.xlsx]Sheet'!$A$1 (cross-file reference)
      - 'Sheet Name'!A1 (cross-sheet reference)
      - Sheet!A1 (same-file cross-sheet)
      - plain A1 or $A$1 (will be prefixed with current_sheet if provided)
    """
    # Match cell references with proper context awareness
    # This pattern matches:
    # 1. Cross-file: '[File.xlsx]Sheet'!A1 or '[File.xlsx]'Sheet'!A1
    # 2. Cross-sheet: 'Sheet Name'!A1 or Sheet!A1
    # 3. Plain cells: A1 or $A$1
    pattern = r"(?:'?\[[^\]]+\]'?)?(?:'[^']+'|[A-Za-z0-9_ ]+)?!\$?[A-Z]{1,3}\$?\d+|\$?[A-Z]{1,3}\$?\d+"

    def replacer(match):
        token = match.group(0)
        ref = token
        
        # Check if this is already a qualified reference (contains '!')
        if '!' in token:
            # It's already qualified (cross-file or cross-sheet), use as-is
            ref = token
            print(f"  📎 Qualified reference: {token}")
        elif current_sheet:
            # Plain cell reference - prefix with current_sheet
            ref = f"{current_sheet}!{token}"
            print(f"  📎 Plain cell {token} -> prefixed as {ref}")
        else:
            # Plain cell but no current_sheet - can't resolve
            print(f"  ⚠️ Plain cell {token} but no current_sheet")
            return token
        
        # Try to resolve the reference
        value = resolve_cell_reference(ref, session_id)
        if value is not None:
            print(f"  ✅ Resolved {ref} = {value}")
            # Keep strings quoted so subsequent parsing works
            if isinstance(value, str):
                return f'"{value}"'
            return str(value)
        
        # If not resolvable, return original token unchanged
        print(f"  ❌ Could not resolve {ref}")
        return token

    print(f"🔍 Resolving references in: {formula}")
    resolved = re.sub(pattern, replacer, formula)
    print(f"🔍 After resolution: {resolved}")
    logger.debug(f"resolve_all_cell_references: '{formula}' -> '{resolved}'")
    # Check if resolved formula contains SUM or AVERAGE
    if 'SUM(' in resolved.upper():
        # Extract and evaluate SUM
        sum_pattern = r'SUM\([^)]+\)'
        def replace_sum(match):
            result = evaluate_sum_function(match.group(0), session_id, current_sheet)
            return str(result) if result is not None else match.group(0)
        resolved = re.sub(sum_pattern, replace_sum, resolved, flags=re.IGNORECASE)

    if 'AVERAGE(' in resolved.upper():
        # Extract and evaluate AVERAGE
        avg_pattern = r'AVERAGE\([^)]+\)'
        def replace_avg(match):
            result = evaluate_average_function(match.group(0), session_id, current_sheet)
            return str(result) if result is not None else match.group(0)
        resolved = re.sub(avg_pattern, replace_avg, resolved, flags=re.IGNORECASE)
        
    # Handle ROUNDUP
    if 'ROUNDUP(' in resolved.upper():
        def replace_roundup(match):
            result = evaluate_roundup_function(match.group(0), session_id, current_sheet)
            return str(result) if result is not None else match.group(0)
        resolved = re.sub(r'ROUNDUP\((?:[^()]+|\([^()]*\))*\)', replace_roundup, resolved, flags=re.IGNORECASE)
        
    # Handle ROUND (with nested parentheses support)
    if 'ROUND(' in resolved.upper():
        def replace_round(match):
            result = evaluate_round_function(match.group(0), session_id, current_sheet)
            return str(result) if result is not None else match.group(0)
        resolved = re.sub(r'ROUND\((?:[^()]+|\([^()]*\))*\)', replace_round, resolved, flags=re.IGNORECASE)

    # Handle SQRT (with nested parentheses support)
    if 'SQRT(' in resolved.upper():
        def replace_sqrt(match):
            result = evaluate_sqrt_function(match.group(0), session_id, current_sheet)
            return str(result) if result is not None else match.group(0)
        resolved = re.sub(r'SQRT\((?:[^()]+|\([^()]*\))*\)', replace_sqrt, resolved, flags=re.IGNORECASE)
        
    # Handle IFERROR (with nested parentheses support)
    if 'IFERROR(' in resolved.upper():
        def replace_iferror(match):
            result = evaluate_iferror_function(match.group(0), session_id, current_sheet)
            return str(result) if result is not None else match.group(0)
        resolved = re.sub(r'IFERROR\((?:[^()]+|\((?:[^()]+|\([^()]*\))*\))*\)', replace_iferror, resolved, flags=re.IGNORECASE)
        print(f"🔍 After IFERROR replacement: {resolved}")
        
    # Handle IF functions embedded in expressions
    if 'IF(' in resolved.upper():
        def replace_if(match):
            if_formula = match.group(0)
            print(f"  🔢 Found nested IF function: {if_formula}")
            result = evaluate_if_function(if_formula, session_id, current_sheet)
            return str(result) if result is not None else if_formula
        # Pattern to match IF with nested content
        resolved = re.sub(r'IF\((?:[^()]+|\((?:[^()]+|\([^()]*\))*\))*\)', replace_if, resolved, flags=re.IGNORECASE)
        print(f"🔍 After IF replacement: {resolved}")

    # Handle OR functions embedded in expressions  
    if 'OR(' in resolved.upper():
        def replace_or(match):
            or_formula = match.group(0)
            print(f"  🔢 Found nested OR function: {or_formula}")
            result = evaluate_or_function(or_formula, session_id, current_sheet)
            return str(result) if result is not None else or_formula
        resolved = re.sub(r'OR\((?:[^()]+|\([^()]*\))*\)', replace_or, resolved, flags=re.IGNORECASE)
        print(f"🔍 After OR replacement: {resolved}")


    return resolved


def safe_eval(expression):
    """Safely evaluate a mathematical expression"""
    try:
        # Remove any quotes around strings
        expression = expression.strip()
        
        # Convert Excel exponentiation (^) to Python (**)
        expression = expression.replace('^', '**')
        
        # Don't process if it contains Excel functions
        if any(func in expression.upper() for func in ['IF(', 'OR(', 'AND(']):
            return None
        
        # Remove quotes from the expression for numeric evaluation
        expression = expression.replace('"', '')
        
        logger.debug(f"safe_eval input: {expression}")
        print(f"🔢 Evaluating: {expression}")
        
        # Only allow basic math operations and numbers
        allowed_chars = set('0123456789+-*/()., ')
        if not all(c in allowed_chars or c.isspace() for c in expression):
            logger.warning(f"Expression contains invalid characters: {expression}")
            print("⚠️ Invalid characters in expression")
            return None
        
        # Evaluate
        result = eval(expression, {"__builtins__": {}}, {})
        logger.debug(f"safe_eval result: {result}")
        print(f"🔢 Result: {result}")
        return result
    except Exception as e:
        logger.error(f"Error in safe_eval for expression '{expression}': {str(e)}", exc_info=True)
        print(f"❌ Evaluation error: {str(e)}")
        return None


def extract_formulas_from_sheet(ws, workbook_path):
    """
    Extract both formulas and normal cell values.
    Adds fields: formula, value, and is_formula.
    Replaces short references like [5], [6], [7], [8] with actual workbook names.
    Handles Array Formulas properly.
    """
    workbook_map = {
        "5": "Pavement Input",
        "6": "TCS Schedule",
        "7": "TCS Input",
        "8": "Emb Height"
    }

    extracted = {}

    for row in ws.iter_rows():
        for cell in row:
            cell_info = {"formula": None, "value": None, "is_formula": False}

            # Check if cell contains a formula
            if cell.data_type == "f" or (cell.value and str(cell.value).startswith("=")):
                formula_value = None
                
                # Handle Array Formula objects
                if hasattr(cell.value, 'text'):
                    # It's an ArrayFormula object, get the text attribute
                    formula_value = cell.value.text
                    logger.debug(f"Extracted array formula from {cell.coordinate}: {formula_value}")
                    print(f"  📋 Array formula in {cell.coordinate}")
                elif isinstance(cell.value, str):
                    # It's a regular string formula
                    formula_value = cell.value
                else:
                    # Try to convert to string as fallback
                    try:
                        formula_value = str(cell.value)
                        # Check if it's the object representation (contains "object at")
                        if "object at" in formula_value:
                            logger.warning(f"Cell {cell.coordinate} has unhandled formula type: {type(cell.value)}")
                            print(f"  ⚠️ Warning: Unhandled formula type in {cell.coordinate}")
                            continue
                    except Exception as e:
                        logger.error(f"Error converting formula in {cell.coordinate}: {e}")
                        continue

                if formula_value:
                    # Replace short workbook references
                    for key, workbook_name in workbook_map.items():
                        pattern = rf"'?\[{key}\]'?('?)([A-Za-z0-9 _-]+)\1'?!"

                        def replace_ref(match):
                            sheet_name = match.group(2).strip()
                            return f"'[{workbook_name}.xlsx]{sheet_name}'!"

                        formula_value = re.sub(pattern, replace_ref, formula_value)

                    # Cleanup malformed quotes
                    formula_value = re.sub(r"'+!'+", "'!", formula_value)
                    formula_value = re.sub(r"!'+!", "!'", formula_value)
                    formula_value = re.sub(r"''", "'", formula_value)

                    cell_info["formula"] = formula_value
                    cell_info["is_formula"] = True

            # If not a formula, store value
            elif cell.value is not None:
                cell_info["value"] = cell.value
                cell_info["is_formula"] = False

            if cell_info["formula"] or cell_info["value"] is not None:
                extracted[cell.coordinate] = cell_info

    logger.info(f"Extracted {len(extracted)} cells (formulas + values) from '{ws.title}'.")
    return extracted

def get_completed_rows_from_mongodb(session_id, sheet_name):
    """Query MongoDB for all completed rows for this session+sheet"""
    try:
        query = {
            "session_id": session_id,
            "sheet_name": sheet_name
        }
        
        completed_docs = list(calculated_main_carriageway_collection.find(query))
        
        completed_rows = {
            doc["row_number"]: doc 
            for doc in completed_docs
        }
        
        return completed_rows
    except Exception as e:
        logger.error(f"Error getting completed rows: {e}")
        return {}


def filter_remaining_rows(all_rows, completed_rows_dict):
    """Filter out completed rows, return only rows needing calculation"""
    remaining = []
    
    for row_num in all_rows:
        if row_num not in completed_rows_dict:
            remaining.append(row_num)
    
    return remaining


def update_summary_from_existing(completed_rows_dict, overall_summary):
    """Update counters from already-completed rows"""
    for row_num, row_doc in sorted(completed_rows_dict.items()):
        print(f"\n{'='*60}")
        print(f"⏭️  Row {row_num} ALREADY CALCULATED - SKIPPING")
        print(f"{'='*60}")
        logger.info(f"Row {row_num} already exists in database, skipping calculation")
        
        overall_summary["total_rows_processed"] += 1
        overall_summary["total_cells_processed"] += row_doc.get("cells_in_row", 0)
        overall_summary["successful_cells"] += row_doc.get("successful_calculations", 0)
        overall_summary["failed_cells"] += row_doc.get("failed_calculations", 0)
        overall_summary["rows_saved"] += 1
        
        print(f"  ℹ️  Using existing calculation from: {row_doc.get('timestamp')}")
        print(f"  ✅ Successful: {row_doc.get('successful_calculations', 0)} cells")
        print(f"  ❌ Failed: {row_doc.get('failed_calculations', 0)} cells")
    
    return len(completed_rows_dict)


def process_single_row(row_num, row_cells, session_id, sheet_name, calculation_id):
    """Process a single row - pure function for thread pool"""
    row_results = []
    row_errors = []
    
    print(f"\n{'='*60}")
    print(f"🔢 Processing Row {row_num} ({len(row_cells)} cells)")
    print(f"{'='*60}")
    logger.info(f"Starting calculation for Row {row_num} with {len(row_cells)} cells")
    
    for formula_doc in row_cells:
        try:
            cell = formula_doc.get("cell")
            sheet = formula_doc.get("sheet")
            is_formula = formula_doc.get("is_formula")
            
            if not cell:
                row_errors.append({
                    "cell": "unknown",
                    "error": "Missing cell in database document"
                })
                continue
            
            if is_formula:
                formula = formula_doc.get("formula")
                if not formula:
                    row_errors.append({
                        "cell": cell,
                        "sheet": sheet,
                        "error": "Formula field is null but is_formula is true"
                    })
                    continue
                
                print(f"  📝 Calculating {sheet}!{cell}: {formula[:50]}...")
                value = evaluate_excel_formula(formula, session_id, current_sheet=sheet)
                
                row_results.append({
                    "cell": cell,
                    "sheet": sheet,
                    "row_number": row_num,
                    "is_formula": True,
                    "formula": formula,
                    "value": value,
                    "success": value is not None
                })
                
                if value is not None:
                    print(f"  ✅ {cell} = {value}")
                else:
                    print(f"  ❌ {cell} = None (calculation failed)")
            else:
                value = formula_doc.get("value")
                
                row_results.append({
                    "cell": cell,
                    "sheet": sheet,
                    "row_number": row_num,
                    "is_formula": False,
                    "formula": None,
                    "value": value,
                    "success": True
                })
                
                print(f"  📌 {cell} = {value} (direct value)")
                
        except Exception as calc_error:
            row_errors.append({
                "cell": cell or "unknown",
                "sheet": sheet,
                "error": str(calc_error)
            })
            logger.error(f"Error calculating cell {cell}: {str(calc_error)}", exc_info=True)
            print(f"  ❌ Error in {cell}: {str(calc_error)}")
    
    row_doc = {
        "calculation_id": calculation_id,
        "session_id": session_id,
        "sheet_name": sheet_name,
        "row_number": row_num,
        "timestamp": datetime.now(timezone.utc),
        "cells_in_row": len(row_cells),
        "successful_calculations": len([r for r in row_results if r["success"]]),
        "failed_calculations": len(row_errors),
        "results": row_results,
        "errors": row_errors
    }
    
    return row_doc


def update_progress_in_redis(session_id, completed, total):
    """Update progress in Redis for tracking"""
    if redis_client is None:
        return
    
    try:
        percent = (completed / total * 100) if total > 0 else 0
        progress_data = {
            "total_rows": total,
            "completed_rows": completed,
            "percent": round(percent, 2)
        }
        
        redis_client.setex(
            f"progress:{session_id}",
            3600,  # 1 hour TTL
            str(progress_data)
        )
    except Exception as e:
        logger.error(f"Error updating progress: {e}")


@app.route("/api/upload-boq-template", methods=["POST"])
def upload_boq_template():
    """Upload BOQ template and identify the BOQ sheet"""
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "Empty filename"}), 400

        if not allowed_file(file.filename):
            return jsonify({"error": "Invalid file type"}), 400

        template_id = str(uuid.uuid4())
        filename = secure_filename(file.filename)
        file_extension = os.path.splitext(filename)[1]
        
        temp_filepath = os.path.join(UPLOAD_FOLDER, f"boq_template_{template_id}{file_extension}")
        file.save(temp_filepath)

        # Load workbook
        wb = openpyxl.load_workbook(temp_filepath, data_only=False)
        
        # Find BOQ sheet
        identified_sheet = None
        for sheet_name in wb.sheetnames:
            if 'BOQ' in sheet_name.upper():
                identified_sheet = sheet_name
                break
        
        if not identified_sheet:
            identified_sheet = wb.sheetnames[0]
        
        ws = wb[identified_sheet]
        
        # Extract BOQ items
        boq_items = []
        headers = []
        
        # Find header row (usually row 1)
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value))
            else:
                headers.append(f"Column_{col}")
        
        # Extract data rows
        for row in range(2, ws.max_row + 1):
            row_data = {}
            has_data = False
            
            for col in range(1, len(headers) + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data[headers[col-1]] = cell_value
                if cell_value:
                    has_data = True
            
            if has_data:
                boq_items.append(row_data)
        
        # Save to MongoDB
        template_doc = {
            "_id": template_id,
            "filename": filename,
            "file_extension": file_extension,
            "identified_sheet": identified_sheet,
            "headers": headers,
            "boq_items": boq_items,
            "uploaded_at": datetime.now(timezone.utc),
            "item_count": len(boq_items)
        }
        
        boq_templates_collection.insert_one(template_doc)

        # Save original file
        output_path = os.path.join(OUTPUT_FOLDER, f"original_boq_template{file_extension}")
        file.seek(0)
        with open(output_path, 'wb') as f:
            f.write(file.read())

        print(f"✅ BOQ Template uploaded: {filename}")
        print(f"✅ Identified sheet: {identified_sheet}")
        print(f"✅ BOQ items extracted: {len(boq_items)}")

        return jsonify({
            "message": "BOQ template uploaded successfully",
            "template_id": template_id,
            "filename": filename,
            "identified_sheet": identified_sheet,
            "item_count": len(boq_items),
            "headers": headers
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/extract-main-carriageway-formulas", methods=["POST"])
def extract_main_carriageway_formulas():
    """Extract formulas and values from Main Carriageway template and save to MongoDB"""
    try:
        # Check if main carriageway template exists
        template_path = os.path.join(OUTPUT_FOLDER, "main_carriageway_template.xlsx")
        if not os.path.exists(template_path):
            return jsonify({"error": "Main Carriageway template not found. Please upload it first."}), 404

        # Load workbook with formulas
        wb = openpyxl.load_workbook(template_path, data_only=False)
        logger.info(f"Workbook 'Main Carriageway.xlsx' loaded with sheets: {wb.sheetnames}")
        
        total_saved = 0
        processed_sheets = []
        
        # Extract formulas and values from all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Use the new extraction function
            formulas = extract_formulas_from_sheet(ws, template_path)
            
            # Prepare documents for MongoDB
            documents = [
                {
                    "file_name": "Main Carriageway.xlsx",
                    "sheet": sheet_name,
                    "cell": cell,
                    "formula": data.get("formula"),
                    "value": data.get("value"),
                    "is_formula": data.get("is_formula", False),
                    "uploaded_at": datetime.now(timezone.utc),
                }
                for cell, data in formulas.items()
            ]

            if documents:
                # Clear existing data for this sheet before inserting
                main_carriageway_formulas_collection.delete_many({
                    "file_name": "Main Carriageway.xlsx",
                    "sheet": sheet_name
                })
                
                # Insert new data
                result = main_carriageway_formulas_collection.insert_many(documents)
                inserted_count = len(result.inserted_ids)
                total_saved += inserted_count
                logger.info(f"Inserted {inserted_count} cells (formulas + values) from '{sheet_name}' into MongoDB.")
                
                processed_sheets.append({
                    "name": sheet_name,
                    "cell_count": inserted_count
                })

        logger.info(f"Total formulas and values saved to MongoDB: {total_saved}")
        print(f"✅ Total cells extracted: {total_saved}")

        return jsonify({
            "message": "Formulas and values extracted successfully",
            "total_saved": total_saved,
            "sheets": processed_sheets
        }), 200

    except Exception as e:
        logger.error(f"Error extracting formulas: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/api/upload-main-carriageway-template", methods=["POST"])
def upload_main_carriageway_template():
    """Upload Main Carriageway template"""
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "Empty filename"}), 400

        if not allowed_file(file.filename):
            return jsonify({"error": "Invalid file type"}), 400

        filename = secure_filename(file.filename)
        file_extension = os.path.splitext(filename)[1]

        # Save to uploads folder (source template)
        upload_path = os.path.join(UPLOAD_FOLDER, f"main_carriageway_template{file_extension}")
        file.seek(0)
        with open(upload_path, 'wb') as f:
            f.write(file.read())

        # Save to outputs folder (for modifications)
        output_path = os.path.join(OUTPUT_FOLDER, f"main_carriageway_template{file_extension}")
        file.seek(0)
        with open(output_path, 'wb') as f:
            f.write(file.read())

        logger.info(f"Main Carriageway Template uploaded: {filename}")
        print(f"✅ Main Carriageway Template uploaded: {filename}")
        print("✅ Saved to uploads and outputs folders")

        # Load workbook to get sheet names for response
        wb = openpyxl.load_workbook(upload_path)
        
        return jsonify({
            "message": "Main Carriageway template uploaded successfully",
            "filename": filename,
            "sheets": wb.sheetnames
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/upload-input-files", methods=["POST"])
def upload_input_files():
    """Upload four input files and save cell values to MongoDB"""
    try:
        # Generate session ID
        session_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        
        # File mappings
        file_mappings = {
            "pavement_input": {
                "collection": pavement_input_values_collection,
                "filename": "Pavement Input.xlsx"
            },
            "tcs_input": {
                "collection": tcs_input_values_collection,
                "filename": "TCS Input.xlsx"
            },
            "emb_height": {
                "collection": emb_height_values_collection,
                "filename": "Emb Height.xlsx"
            },
            "tcs_schedule": {
                "collection": tcs_schedule_values_collection,
                "filename": "TCS Schedule.xlsx"
            }
        }
        
        uploaded_files = {}
        total_cells = 0
        
        for file_key, mapping in file_mappings.items():
            if file_key not in request.files:
                return jsonify({"error": f"Missing file: {file_key}"}), 400
            
            file = request.files[file_key]
            if file.filename == "":
                return jsonify({"error": f"Empty filename for {file_key}"}), 400
            
            if not allowed_file(file.filename):
                return jsonify({"error": f"Invalid file type for {file_key}"}), 400
            
            # Save temporarily
            temp_filepath = os.path.join(UPLOAD_FOLDER, f"{file_key}_{session_id}.xlsx")
            file.save(temp_filepath)
            
            # Load workbook
            wb = openpyxl.load_workbook(temp_filepath, data_only=True)
            
            collection = mapping["collection"]
            filename = mapping["filename"]
            cell_count = 0
            
            # Extract all cell values from all sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell_doc = {
                                "session_id": session_id,
                                "file_name": filename,
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "value": cell.value,
                                "uploaded_at": datetime.now(timezone.utc)
                            }
                            
                            collection.insert_one(cell_doc)
                            cell_count += 1
            
            uploaded_files[file_key] = {
                "filename": filename,
                "sheets": wb.sheetnames,
                "cell_count": cell_count
            }
            
            total_cells += cell_count
        
        # Create file session document
        session_doc = {
            "session_id": session_id,  # Store session_id as a field
            "created_at": datetime.now(timezone.utc),
            "type": "file_upload",
            "uploaded_files": uploaded_files,
            "total_cells": total_cells
        }
        
        # Insert and get the MongoDB generated _id
        result = file_sessions_collection.insert_one(session_doc)
        mongo_id = result.inserted_id

        logger.info(f"Created file upload session with ID: {session_id}")
        print(f"✅ Session created: {session_id}")
        print(f"✅ Total cells stored: {total_cells}")
        print(f"✅ MongoDB document ID: {mongo_id}")

        return jsonify({
            "message": "Input files uploaded successfully",
            "session_id": session_id,
            "mongo_id": str(mongo_id),
            "uploaded_files": uploaded_files,
            "total_cells": total_cells
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/calculate-boq", methods=["POST"])
def calculate_boq():
    """Calculate BOQ values by evaluating Main Carriageway formulas"""
    try:
        logger.info("Starting BOQ calculation")
        data = request.json
        session_id = data.get("session_id")
        
        if not session_id:
            logger.error("Missing session_id in request")
            return jsonify({"error": "session_id is required"}), 400

        # Verify file upload session exists
        session = file_sessions_collection.find_one({"session_id": session_id})
        if not session:
            return jsonify({"error": "Session not found"}), 404
        
        # Get BOQ template
        boq_template = boq_templates_collection.find_one(sort=[("uploaded_at", -1)])
        if not boq_template:
            return jsonify({"error": "No BOQ template found"}), 404
        
        boq_items = boq_template.get("boq_items", [])
        
        # Removed unused query for Main Carriageway formulas
        
        calculated_results = []
        
        # For each BOQ item, find corresponding formula and calculate
        for idx, item in enumerate(boq_items):
            # Find sub_bill_id or item identifier
            sub_bill_id = None
            for key, value in item.items():
                if value and (isinstance(value, str) or isinstance(value, (int, float))):
                    sub_bill_id = str(value)
                    break
            
            if not sub_bill_id:
                continue
            
            # Try to find matching formula in Abstract sheet
            # Usually Abstract sheet formulas are in sequential rows
            row_num = idx + 2  # Assuming row 1 is headers, data starts at row 2
            
            # Look for formulas in common result columns (like D, E, F, etc.)
            main_carriageway_value = None
            service_road_value = None
            calculation_notes = ""
            
            # Try to find formula for this row
            for col_letter in ['D', 'E', 'F', 'G', 'H']:  # Common result columns
                cell_address = f"{col_letter}{row_num}"
                
                formula_doc = main_carriageway_formulas_collection.find_one({
                    "file_name": "Main Carriageway.xlsx",
                    "sheet": "Abstract",
                    "cell": cell_address
                })
                
                if formula_doc:
                    formula = formula_doc.get("formula")
                    if formula:
                        result = evaluate_excel_formula(formula, session_id, current_sheet=formula_doc.get("sheet"))
                        
                        if main_carriageway_value is None:
                            main_carriageway_value = result
                            calculation_notes += f"Cell {cell_address}: {formula} = {result}; "
                        elif service_road_value is None:
                            service_road_value = result
                            calculation_notes += f"Cell {cell_address}: {formula} = {result}; "
            
            result_doc = {
                "sub_bill_id": sub_bill_id,
                "item": item,
                "main_carriageway": main_carriageway_value,
                "service_road": service_road_value,
                "calculation_notes": calculation_notes.strip()
            }
            
            calculated_results.append(result_doc)
        
        # Update session with results
        file_sessions_collection.update_one(
            {"session_id": session_id},
            {"$set": {
                "calculated_results": calculated_results,
                "calculated_at": datetime.now(timezone.utc),
                "status": "calculated"  # Add status to track calculation state
            }}
        )
        
        print(f"✅ Calculated {len(calculated_results)} items")

        return jsonify({
            "message": "Calculation completed",
            "session_id": session_id,
            "calculated_items": len(calculated_results),
            "results": calculated_results[:10]  # Return first 10 for preview
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/save-in-boq-template", methods=["POST"])
def save_in_boq_template():
    """Update the BOQ template Excel file with calculated values"""
    try:
        data = request.json
        session_id = data.get("session_id")

        if not session_id:
            return jsonify({"error": "session_id is required"}), 400

        # Get session data
        session = file_sessions_collection.find_one({"session_id": session_id})
        if not session:
            return jsonify({"error": "Session not found"}), 404

        calculated_results = session.get("calculated_results", [])
        if not calculated_results:
            return jsonify({"error": "No calculated results found. Run calculations first."}), 400

        # Get BOQ template
        boq_template = boq_templates_collection.find_one(sort=[("uploaded_at", -1)])
        if not boq_template:
            return jsonify({"error": "BOQ template not found"}), 404

        identified_sheet = boq_template["identified_sheet"]
        file_extension = boq_template["file_extension"]
        # original_filename not needed, skipping

        # Load original BOQ file
        original_boq_path = os.path.join(OUTPUT_FOLDER, f"original_boq_template{file_extension}")

        if not os.path.exists(original_boq_path):
            return jsonify({"error": "Original BOQ Excel file not found"}), 404

        wb = openpyxl.load_workbook(original_boq_path)
        
        if identified_sheet not in wb.sheetnames:
            return jsonify({"error": f"Sheet '{identified_sheet}' not found"}), 404
        
        ws = wb[identified_sheet]

        # Find or create columns
        header_row = 1
        headers = {}

        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value:
                headers[str(cell_value).lower()] = col_idx

        # Find or create result columns
        main_col = None
        service_col = None
        notes_col = None

        for header_text, col_idx in headers.items():
            if "main" in header_text and "carriageway" in header_text:
                main_col = col_idx
            if "service" in header_text and "road" in header_text:
                service_col = col_idx
            if "calculation" in header_text and "note" in header_text:
                notes_col = col_idx

        if main_col is None:
            main_col = ws.max_column + 1
            ws.cell(row=header_row, column=main_col).value = "Main Carriageway"

        if service_col is None:
            service_col = ws.max_column + 1
            ws.cell(row=header_row, column=service_col).value = "Service Road"

        if notes_col is None:
            notes_col = ws.max_column + 1
            ws.cell(row=header_row, column=notes_col).value = "Calculation Notes"

        # Update rows with calculated values
        updated_count = 0

        for idx, result in enumerate(calculated_results):
            row_idx = idx + 2  # Row 1 is header, data starts at row 2
            
            if row_idx <= ws.max_row:
                ws.cell(row=row_idx, column=main_col).value = result.get("main_carriageway")
                ws.cell(row=row_idx, column=service_col).value = result.get("service_road")
                ws.cell(row=row_idx, column=notes_col).value = result.get("calculation_notes")
                updated_count += 1

        # Save files
        wb.save(original_boq_path)
        
        session_output_filename = f"BOQ_Updated_{session_id}{file_extension}"
        session_output_path = os.path.join(OUTPUT_FOLDER, session_output_filename)
        wb.save(session_output_path)

        print(f"✅ Updated {updated_count} rows")

        mime_types = {
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.xlsm': 'application/vnd.ms-excel.sheet.macroEnabled.12',
            '.xlsb': 'application/vnd.ms-excel.sheet.binary.macroEnabled.12',
        }

        mime_type = mime_types.get(file_extension, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        return send_file(
            session_output_path,
            as_attachment=True,
            download_name=session_output_filename,
            mimetype=mime_type,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/api/sessions", methods=["GET"])
def get_all_sessions():
    """Get all file upload sessions with pagination (lite version)"""
    try:
        # Get query parameters
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        
        # Calculate skip for pagination
        skip = (page - 1) * limit
        
        # Get sessions with pagination and sorting (newest first)
        sessions = list(file_sessions_collection.find(
            {},
            {
                "_id": 1,
                "session_id": 1,
                "created_at": 1,
                "calculated_at": 1,
                "uploaded_files": 1,
                "calculated_results": 1
            }
        ).sort("created_at", -1).skip(skip).limit(limit))
        
        # Convert to JSON-serializable format
        serialized_sessions = []
        for session in sessions:
            # Determine status
            calculated_results = session.get("calculated_results", [])
            if not calculated_results:
                status = "not calculated"
            elif calculated_results and len(calculated_results) > 0:
                status = "calculated"
            else:
                status = "partially calculated"
            
            # Build file summary
            file_summary = {}
            uploaded_files = session.get("uploaded_files", {})
            for file_key, file_info in uploaded_files.items():
                file_summary[file_key] = {
                    "filename": file_info.get("filename"),
                    "sheets": file_info.get("sheets", []),
                    "cell_count": file_info.get("cell_count", 0)
                }
            
            serialized_session = {
                "_id": str(session["_id"]),
                "session_id": session.get("session_id"),
                "created_at": session.get("created_at").isoformat() if session.get("created_at") else None,
                "calculated_at": session.get("calculated_at").isoformat() if session.get("calculated_at") else None,
                "status": status,
                "file_summary": file_summary
            }
            
            serialized_sessions.append(serialized_session)
        
        logger.info(f"Retrieved {len(serialized_sessions)} file sessions")
        
        return jsonify(serialized_sessions), 200

    except Exception as e:
        logger.error(f"Error retrieving sessions: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/sessions/<session_id>", methods=["GET"])
def get_session(session_id):
    """Get session details"""
    try:
        # Try to find in app sessions first
        session = file_sessions_collection.find_one({"session_id": session_id})
        if not session:
            # If not found, try file sessions
            session = file_sessions_collection.find_one({"_id": session_id})
        
        if not session:
            return jsonify({"error": "Session not found"}), 404
        
        # Convert to JSON-serializable format
        if "_id" in session:
            session["_id"] = str(session["_id"])
            
        logger.info(f"Retrieved {session.get('type', 'unknown')} session: {session_id}")
        
        return jsonify(session), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/calculate-main-carriageway", methods=["POST"])
def calculate_main_carriageway():
    """Calculate values using Main Carriageway formulas - PARALLEL ROW-BY-ROW with Resume"""
    try:
        data = request.json
        session_id = data.get("session_id")
        sheet_name = data.get("sheet_name")
        max_workers = data.get("max_workers", 50)  # Allow customization, default 50
        calculation_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")

        if not session_id:
            return jsonify({"error": "Session ID is required"}), 400
        
        if not sheet_name:
            return jsonify({"error": "Sheet name is required"}), 400

        # Verify file upload session exists
        session = file_sessions_collection.find_one({"session_id": session_id})
        if not session:
            return jsonify({"error": f"No file upload session found with ID: {session_id}"}), 404

        # ✅ NEW: Load input data to memory for this session
        logger.info(f"Loading input data to memory for session: {session_id}")
        print(f"🔄 Loading input data to memory for session: {session_id}")
        load_success = load_input_data_to_memory(session_id)

        if load_success:
            logger.info("Input data loaded successfully to memory")
            print("✅ Input data loaded successfully to memory")
        else:
            logger.warning("Input data load to memory failed, will use MongoDB directly")
            print("⚠️ Input data load to memory failed, using MongoDB directly")
            
        # Load Main Carriageway formulas to memory
        logger.info("Loading Main Carriageway formulas to memory")
        print("🔄 Loading Main Carriageway formulas to memory...")
        formula_load_success = load_main_carriageway_formulas_to_memory(sheet_name)
        
        if formula_load_success:
            logger.info("Main Carriageway formulas loaded successfully to memory")
            print("✅ Main Carriageway formulas loaded to memory")
        else:
            logger.warning("Formula load failed, will query MongoDB directly")
            print("⚠️ Formula load failed, using MongoDB directly")

        # Get formulas from database for the specified sheet
        formulas = list(main_carriageway_formulas_collection.find({"sheet": sheet_name}))
        if not formulas:
            return jsonify({"error": f"No formulas found for sheet '{sheet_name}' in database"}), 404
        
        logger.info(f"Found {len(formulas)} formulas for sheet '{sheet_name}'")
        print(f"📄 Sheet: {sheet_name}")

        # Group cells by row number
        rows_dict = {}
        for formula_doc in formulas:
            cell = formula_doc.get("cell")
            if not cell:
                continue
            
            import re
            row_match = re.search(r'\d+', cell)
            if row_match:
                row_num = int(row_match.group())
                if row_num not in rows_dict:
                    rows_dict[row_num] = []
                rows_dict[row_num].append(formula_doc)
        
        sorted_rows = sorted(rows_dict.keys())
        
        logger.info(f"Processing sheet '{sheet_name}': {len(sorted_rows)} rows with {len(formulas)} total cells")
        print(f"📄 Sheet: {sheet_name}")
        print(f"📊 Total rows to process: {len(sorted_rows)}")
        print(f"📊 Total cells to process: {len(formulas)}")
        
        # ═══════════════════════════════════════════════════════════
        # PHASE 1: PRE-FILTER (Main Thread - Sequential)
        # ═══════════════════════════════════════════════════════════
        
        print(f"\n{'='*60}")
        print("🔍 PHASE 1: Checking for existing calculations...")
        print(f"{'='*60}")
        
        # Get completed rows from MongoDB
        completed_rows_dict = get_completed_rows_from_mongodb(session_id, sheet_name)
        
        # Filter remaining rows
        remaining_rows = filter_remaining_rows(sorted_rows, completed_rows_dict)
        
        # Initialize summary
        overall_summary = {
            "total_rows_processed": 0,
            "total_cells_processed": 0,
            "successful_cells": 0,
            "failed_cells": 0,
            "rows_saved": 0
        }
        
        # Update summary from existing rows
        initial_progress = update_summary_from_existing(completed_rows_dict, overall_summary)
        
        print(f"\n{'='*60}")
        print("📊 PRE-FILTER SUMMARY")
        print(f"{'='*60}")
        print(f"Total rows in sheet: {len(sorted_rows)}")
        print(f"Already completed: {len(completed_rows_dict)}")
        print(f"Remaining to process: {len(remaining_rows)}")
        print(f"{'='*60}\n")
        
        # If all rows already completed, return immediately
        if len(remaining_rows) == 0:
            print("✅ All rows already calculated!")
            
            clear_session_cache(session_id)
            
            response = {
                "calculation_id": calculation_id,
                "session_id": session_id,
                "sheet_name": sheet_name,
                "message": f"All rows already calculated for sheet '{sheet_name}'",
                "summary": overall_summary
            }
            
            return jsonify(response), 200
        
        # ═══════════════════════════════════════════════════════════
        # PHASE 2: PARALLEL EXECUTION (Thread Pool)
        # ═══════════════════════════════════════════════════════════
        
        print(f"\n{'='*60}")
        print(f"🚀 PHASE 2: Starting parallel calculation with {max_workers} workers...")
        print(f"{'='*60}\n")
        
        # Thread-safe counter
        counter_lock = Lock()
        completed_count = 0
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all remaining rows to thread pool
            futures = {}
            for row_num in remaining_rows:
                future = executor.submit(
                    process_single_row,
                    row_num,
                    rows_dict[row_num],
                    session_id,
                    sheet_name,
                    calculation_id
                )
                futures[future] = row_num
            
            print(f"📤 Submitted {len(futures)} rows to thread pool\n")
            
            # Process results as they complete
            for future in as_completed(futures):
                try:
                    row_doc = future.result()
                    row_num = futures[future]
                    
                    # Save to MongoDB immediately
                    calculated_main_carriageway_collection.insert_one(row_doc)
                    
                    # Update counters (thread-safe)
                    with counter_lock:
                        completed_count += 1
                        overall_summary["total_rows_processed"] += 1
                        overall_summary["total_cells_processed"] += row_doc["cells_in_row"]
                        overall_summary["successful_cells"] += row_doc["successful_calculations"]
                        overall_summary["failed_cells"] += row_doc["failed_calculations"]
                        overall_summary["rows_saved"] += 1
                        
                        total_completed = initial_progress + completed_count
                        
                        # Update progress in Redis
                        update_progress_in_redis(session_id, total_completed, len(sorted_rows))
                    
                    logger.info(f"Row {row_num} saved to MongoDB. Success: {row_doc['successful_calculations']}, Errors: {row_doc['failed_calculations']}")
                    print(f"💾 Row {row_num} saved to MongoDB")
                    print(f"  ✅ Successful: {row_doc['successful_calculations']} cells")
                    print(f"  ❌ Failed: {row_doc['failed_calculations']} cells")
                    print(f"  📊 Progress: {total_completed}/{len(sorted_rows)} ({total_completed/len(sorted_rows)*100:.1f}%)\n")
                    
                except Exception as e:
                    row_num = futures[future]
                    logger.error(f"Error processing row {row_num}: {e}", exc_info=True)
                    print(f"❌ Error processing row {row_num}: {e}")
        
        # Final summary
        print(f"\n{'='*60}")
        print(f"🎉 CALCULATION COMPLETE")
        print(f"{'='*60}")
        print(f"Total Rows Processed: {overall_summary['total_rows_processed']}")
        print(f"Total Cells Processed: {overall_summary['total_cells_processed']}")
        print(f"Successful Calculations: {overall_summary['successful_cells']}")
        print(f"Failed Calculations: {overall_summary['failed_cells']}")
        print(f"Rows Saved to MongoDB: {overall_summary['rows_saved']}")
        print(f"{'='*60}\n")
        
        logger.info(f"Calculation completed. Summary: {overall_summary}")
        
        # Clear cache for this session after all calculations
        clear_session_cache(session_id)
        logger.info(f"Cache cleared for session {session_id}")
        
        # Clear Main Carriageway formulas from memory
        clear_main_carriageway_formulas_from_memory()
        
        response = {
            "calculation_id": calculation_id,
            "session_id": session_id,
            "sheet_name": sheet_name,
            "message": f"Parallel calculation completed successfully for sheet '{sheet_name}'",
            "max_workers": max_workers,
            "summary": overall_summary
        }
        
        return jsonify(response), 200
        
    except Exception as e:
        error_msg = f"Error calculating Main Carriageway values: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return jsonify({"error": error_msg}), 500


@app.route("/api/calculation-progress/<session_id>", methods=["GET"])
def get_calculation_progress(session_id):
    """Get real-time calculation progress for a session"""
    try:
        if redis_client is None:
            return jsonify({"error": "Redis not available"}), 503
        
        progress_key = f"progress:{session_id}"
        progress_data = redis_client.get(progress_key)
        
        if not progress_data:
            return jsonify({
                "session_id": session_id,
                "status": "not_found",
                "message": "No progress data found for this session"
            }), 404
        
        import ast
        progress = ast.literal_eval(progress_data)
        
        return jsonify({
            "session_id": session_id,
            "status": "in_progress",
            "progress": progress
        }), 200
        
    except Exception as e:
        logger.error(f"Error getting progress: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/calculate-main-carriageway-single-cell", methods=["POST"])
def calculate_main_carriageway_single_cell():
    """Calculate value for a specific cell in Main Carriageway using its cell number"""
    try:
        data = request.json
        session_id = data.get("session_id")
        cell = data.get("cell")  # Example: "A84" or "BY8"
        sheet_name = data.get("sheet_name")  # Sheet name where the cell is located

        if not all([session_id, cell, sheet_name]):
            return jsonify({"error": "session_id, cell, and sheet_name are required"}), 400

        # Verify file upload session exists
        session = file_sessions_collection.find_one({"session_id": session_id})
        if not session:
            return jsonify({"error": f"No file upload session found with ID: {session_id}"}), 404
            
        # Get formula from database for this cell and sheet
        formula_doc = main_carriageway_formulas_collection.find_one({
            "cell": cell,
            "sheet": sheet_name
        })
        
        if not formula_doc:
            return jsonify({"error": f"No document found for cell {cell} in sheet {sheet_name}"}), 404

        # Check if it's a formula or direct value
        if formula_doc.get("is_formula"):
            # It's a formula - evaluate it
            formula = formula_doc.get("formula")
            if not formula:
                return jsonify({"error": f"Formula field is null for cell {cell}"}), 500
            
            # Calculate the value using the formula
            value = evaluate_excel_formula(formula, session_id, current_sheet=sheet_name)
            
            # Format the response
            result = {
                "cell": cell,
                "sheet": sheet_name,
                "is_formula": True,
                "formula": formula,
                "value": value,
                "success": value is not None
            }
        else:
            # It's a direct value - no calculation needed
            value = formula_doc.get("value")
            
            # Format the response
            result = {
                "cell": cell,
                "sheet": sheet_name,
                "is_formula": False,
                "formula": None,
                "value": value,
                "success": True
            }
        
        return jsonify(result)
        
    except Exception as e:
        error_msg = f"Error calculating cell {cell}: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return jsonify({
            "error": error_msg
        }), 500


@app.route("/api/save-in-main-carriageway", methods=["POST"])
def save_in_main_carriageway():
    """Update the Main Carriageway Excel file with calculated values from MongoDB for a specific session and sheet."""
    try:
        data = request.json
        session_id = data.get("session_id")
        sheet_name = data.get("sheet_name")  # Get the specific sheet name to save

        if not session_id:
            return jsonify({"error": "session_id is required"}), 400
        if not sheet_name:
            return jsonify({"error": "sheet_name is required"}), 400

        # Verify file upload session exists
        session = file_sessions_collection.find_one({"session_id": session_id})
        if not session:
            return jsonify({"error": "Session not found"}), 404

        # Load template from uploads folder (READ ONLY)
        template_path = os.path.join(UPLOAD_FOLDER, "main_carriageway_template.xlsx")
        if not os.path.exists(template_path):
            return jsonify({"error": "Main Carriageway template not found in uploads folder"}), 404

        print(f"📁 Loading template for session: {session_id}, sheet: {sheet_name}")
        wb = openpyxl.load_workbook(template_path)

        # Check if the specified sheet exists in the workbook
        if sheet_name not in wb.sheetnames:
            return jsonify({"error": f"Sheet '{sheet_name}' not found in the Main Carriageway template."}), 404

        ws = wb[sheet_name]

        # OPTIMIZATION: Fast clearing from row 7 onwards
        print("🧹 Fast clearing cells from row 7 onwards...")
        
        # Get all merged cell ranges
        merged_ranges = list(ws.merged_cells.ranges)
        merged_cell_coords = set()
        for merged_range in merged_ranges:
            for row in merged_range:
                for cell in row:
                    merged_cell_coords.add(cell.coordinate)

        # OPTIMIZATION: Use batch clearing - only clear non-merged cells
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row > 6 and max_col > 0:
            cleared_cells = 0
            # Process in batches to show progress
            batch_size = 100  # Process 100 rows at a time for progress reporting
            
            for start_row in range(7, max_row + 1, batch_size):
                end_row = min(start_row + batch_size - 1, max_row)
                print(f"🧹 Clearing rows {start_row}-{end_row}...")
                
                for row_num in range(start_row, end_row + 1):
                    for col_num in range(1, max_col + 1):
                        cell_coord = f"{openpyxl.utils.get_column_letter(col_num)}{row_num}"
                        
                        # Skip merged cells to preserve formatting
                        if cell_coord not in merged_cell_coords:
                            cell = ws.cell(row=row_num, column=col_num)
                            # Only clear if cell has a value (faster than clearing all)
                            if cell.value is not None:
                                cell.value = None
                                cleared_cells += 1
            

        # Retrieve all calculated rows for the given session_id and sheet_name from MongoDB
        print(f"📊 Retrieving calculated data from MongoDB for {sheet_name}...")
        calculated_rows = list(calculated_main_carriageway_collection.find({
            "session_id": session_id,
            "sheet_name": sheet_name
        }).sort("row_number", 1))

        if not calculated_rows:
            logger.info(f"No calculated data found for session '{session_id}' and sheet '{sheet_name}'.")
            # Save the cleared template to OUTPUT folder only
            session_output_filename = f"Main_Carriageway_Updated_{session_id}.xlsx"  # Removed sheet_name from filename
            session_output_path = os.path.join(OUTPUT_FOLDER, session_output_filename)
            wb.save(session_output_path)
            print(f"💾 Saved cleared template to: {session_output_path}")
            return jsonify({
                "error": "No calculated data found for this session and sheet",
                "file_path": session_output_filename
            }), 404

        print(f"📈 Processing {len(calculated_rows)} calculated rows...")

        # OPTIMIZATION: Prepare merged range mapping ONCE
        merged_range_top_left_map = {}
        for merged_range in merged_ranges:
            top_left_coord = merged_range.coord.split(':')[0]
            for row in merged_range:
                for cell in row:
                    merged_range_top_left_map[cell.coordinate] = top_left_coord

        # OPTIMIZATION: Batch cell updates with progress tracking
        updated_cells = 0
        skipped_cells = 0
        total_rows = len(calculated_rows)
        
        # Process in smaller batches for better progress reporting
        batch_size = min(50, max(10, total_rows // 10))
        
        print(f"🔄 Writing calculated values in batches of {batch_size} rows...")
        
        for batch_start in range(0, total_rows, batch_size):
            batch_end = min(batch_start + batch_size, total_rows)
            print(f"📝 Processing rows {batch_start + 1}-{batch_end}...")
            
            for i in range(batch_start, batch_end):
                row_doc = calculated_rows[i]
                row_num = row_doc.get("row_number")
                results = row_doc.get("results", [])

                for cell_result in results:
                    cell_address = cell_result.get("cell")
                    calculated_value = cell_result.get("value")

                    # Skip if no value to write
                    if calculated_value is None:
                        skipped_cells += 1
                        continue

                    # Determine the actual cell address to write to (handle merged cells)
                    target_cell_address = merged_range_top_left_map.get(cell_address, cell_address)
                    
                    # Write the calculated value directly
                    try:
                        ws[target_cell_address].value = calculated_value
                        updated_cells += 1
                    except Exception as cell_error:
                        print(f"⚠️ Error writing to cell {target_cell_address}: {cell_error}")
                        skipped_cells += 1

        # Save ONLY to OUTPUT folder with session_id (original filename format)
        session_output_filename = f"Main_Carriageway_Updated_{session_id}.xlsx"  # Original format without sheet_name
        session_output_path = os.path.join(OUTPUT_FOLDER, session_output_filename)
        
        print(f"💾 Saving to output folder: {session_output_filename}")
        wb.save(session_output_path)
        
        logger.info(f"Saved updated template for sheet '{sheet_name}' to {session_output_path}")
        logger.info(f"Updated {updated_cells} cells, skipped {skipped_cells} cells")

        print(f"✅ Successfully saved {updated_cells} cells in {total_rows} rows")
        print(f"✅ File saved to: {session_output_path}")

        return send_file(
            session_output_path,
            as_attachment=True,
            download_name=session_output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        
    except Exception as e:
        print(f"❌ Error saving to Main Carriageway file: {e}")
        traceback.print_exc()
        logger.error(f"Error in save_in_main_carriageway: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/api/redis-stats/<session_id>", methods=["GET"])
def get_redis_stats(session_id):
    """Get statistics about Redis cache for a session"""
    if redis_client is None:
        return jsonify({"error": "Redis not available"}), 503
    
    try:
        # Count calculated formula keys  
        calc_pattern = f"calc:{session_id}:*"
        calc_keys = redis_client.keys(calc_pattern)
        
        return jsonify({
            "session_id": session_id,
            "calculated_formula_keys": len(calc_keys),
        }), 200
        
    except Exception as e:
        logger.error(f"Error getting Redis stats: {e}")
        return jsonify({"error": str(e)}), 500
    

@app.route("/api/memory-cache-stats/<session_id>", methods=["GET"])
def get_memory_cache_stats(session_id):
    """Get statistics about memory cache for a session"""
    try:
        with cache_lock:
            if session_id in input_data_cache:
                session_data = input_data_cache[session_id]
                return jsonify({
                    "session_id": session_id,
                    "cached_cells": len(session_data),
                    "cache_size_bytes": len(str(session_data).encode('utf-8')),
                    "file_types": list(set(key.split(':')[0] for key in session_data.keys()))
                }), 200
            else:
                return jsonify({
                    "session_id": session_id,
                    "cached_cells": 0,
                    "message": "No data cached for this session"
                }), 404
        
    except Exception as e:
        logger.error(f"Error getting memory cache stats: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/", methods=["GET"])
@app.route("/api", methods=["GET"])
def root():
    """Get API information and list of available endpoints"""
    api_info = {
        "name": "BOQ Calculator API",
        "version": "2.1.0",
        "description": "API for calculating Bill of Quantities (BOQ) with Excel formula support",
        "endpoints": [
            {
                "path": "/api",
                "method": "GET",
                "description": "Get API information and list of endpoints"
            },
            {
                "path": "/api/health",
                "method": "GET",
                "description": "Check API and database health status"
            },
            {
                "path": "/api/upload-boq-template",
                "method": "POST",
                "description": "Upload BOQ template Excel file"
            },
            {
                "path": "/api/upload-main-carriageway-template",
                "method": "POST",
                "description": "Upload Main Carriageway template"
            },
            {
                "path": "/api/extract-main-carriageway-formulas",
                "method": "POST",
                "description": "Extract and save formulas from uploaded Main Carriageway template"
            },
            {
                "path": "/api/upload-input-files",
                "method": "POST",
                "description": "Upload input files (TCS, Pavement, Emb Height, TCS Schedule)"
            },
            {
                "path": "/api/calculate-boq",
                "method": "POST",
                "description": "Calculate BOQ values using Main Carriageway formulas"
            },
            {
                "path": "/api/save-in-boq-template",
                "method": "POST",
                "description": "Save calculated values back to BOQ template"
            },
            {
                "path": "/api/calculate-main-carriageway",
                "method": "POST",
                "description": "Calculate values using Main Carriageway formulas"
            },
            {
                "path": "/api/calculation-progress/{session_id}",
                "method": "GET",
                "description": "Get real-time calculation progress for a session"
            },
            {
                "path": "/api/calculate-main-carriageway-single-cell",
                "method": "POST",
                "description": "Calculate value for a specific cell in Main Carriageway. Required parameters: session_id, cell, sheet_name"
            },
            {
                "path": "/api/save-in-main-carriageway",
                "method": "POST",
                "description": "Save calculated values back to Main Carriageway file"
            },
             {
                "path": "/api/sessions",
                "method": "GET",
                "description": "Get all file sessions with pagination",
                "parameters": {
                    "page": "Page number (default: 1)",
                    "limit": "Items per page (default: 50)"
                }
            },
            {
                "path": "/api/sessions/{session_id}",
                "method": "GET",
                "description": "Get details of a specific calculation session"
            },
            {
                "path": "/api/memory-cache-stats/<session_id>",
                "method": "GET", 
                "description": "Get statistics about memory cache for a session",
                "parameters": {
                    "session_id": "Session ID to check memory cache statistics for"
                }
            },
            {
                "path": "/api/redis-stats/<session_id>",
                "method": "GET", 
                "description": "Get statistics about Redis cache for a session",
                "parameters": {
                    "session_id": "Session ID to check Redis cache statistics for"
                }
            },
            {
                "path": "/api/flush-redis-cache",
                "method": "POST",
                # "description": "Flush entire Redis cache database. Requires confirmation: {\"confirm\": true}"
                "description": "Flush entire Redis cache database."
            },
        ],
        "documentation": "For more details on request/response formats, please refer to the API documentation"
    }
    return jsonify(api_info), 200


@app.route("/api/health", methods=["GET"])
def health_check():
    """Check if API and database are running"""
    try:
        # Test MongoDB connection
        mongo_client.admin.command('ping')
        db_status = "connected"
        
        # Count documents in collections
        stats = {
            "main_carriageway_formulas": main_carriageway_formulas_collection.count_documents({}),
            "tcs_input_values": tcs_input_values_collection.count_documents({}),
            "pavement_input_values": pavement_input_values_collection.count_documents({}),
            "emb_height_values": emb_height_values_collection.count_documents({}),
            "tcs_schedule_values": tcs_schedule_values_collection.count_documents({}),
            "app_sessions": app_sessions_collection.count_documents({}),
            "file_sessions": file_sessions_collection.count_documents({}),
            "boq_templates": boq_templates_collection.count_documents({})
        }
    except Exception as e:
        logger.error(f"Error connecting to database: {str(e)}", exc_info=True)
        db_status = "disconnected"
        stats = {}
    
    return jsonify({
        "status": "healthy",
        "service": "BOQ Calculator API with MongoDB",
        "version": "2.1.0",
        "database": db_status,
        "statistics": stats
    }), 200


@app.route("/api/flush-redis-cache", methods=["POST"])
def flush_redis_cache_endpoint():
    """Flush entire Redis cache database"""
    try:
        if redis_client is None:
            return jsonify({
                "error": "Redis is not configured or unavailable"
            }), 503
        
        # Optional: Require confirmation in request body
        # data = request.json or {}
        # confirmation = data.get("confirm", False)
        
        # if not confirmation:
        #     return jsonify({
        #         "error": "Confirmation required",
        #         "message": "Please send {\"confirm\": true} to flush Redis cache"
        #     }), 400
        
        success, message = flush_redis_cache()
        
        if success:
            logger.info("Redis cache flushed via API endpoint")
            return jsonify({
                "message": message,
                "status": "success"
            }), 200
        else:
            return jsonify({
                "error": message,
                "status": "failed"
            }), 500
            
    except Exception as e:
        logger.error(f"Error in flush-redis-cache endpoint: {str(e)}", exc_info=True)
        return jsonify({
            "error": str(e),
            "status": "failed"
        }), 500


if __name__ == "__main__":
    # Log startup information
    logger.info(f"Starting BOQ Calculator API... Session ID: {current_session_id}")
    
    # Check database connection
    try:
        mongo_client.admin.command('ping')
        logger.info("Successfully connected to MongoDB")
        
        # Log collection statistics
        stats = {
            "main_carriageway_formulas": main_carriageway_formulas_collection.count_documents({}),
            "tcs_input_values": tcs_input_values_collection.count_documents({}),
            "pavement_input_values": pavement_input_values_collection.count_documents({}),
            "emb_height_values": emb_height_values_collection.count_documents({}),
            "tcs_schedule_values": tcs_schedule_values_collection.count_documents({}),
            "app_sessions": app_sessions_collection.count_documents({}),
            "file_sessions": file_sessions_collection.count_documents({}),
            "boq_templates": boq_templates_collection.count_documents({})
        }
        logger.info(f"Database statistics: {stats}")
        
        # Save application session info to database
        session_info = {
            "_id": current_session_id,
            "type": "application_session",
            "start_time": datetime.now(),
            "log_file": os.path.join(SESSIONS_LOG_FOLDER, f'session_{current_session_id}.log'),
            "database_stats": stats
        }
        app_sessions_collection.insert_one(session_info)
        logger.info(f"Application session info saved to database with ID: {current_session_id}")
        
    except Exception as e:
        logger.error(f"Failed to connect to MongoDB: {str(e)}", exc_info=True)
        raise SystemExit("Could not establish database connection. Exiting...")
    
    # Log server configuration
    logger.info("Server starting on http://0.0.0.0:5000")
    logger.info(f"Debug mode: {app.debug}")
    logger.info(f"Upload folder: {os.path.abspath(UPLOAD_FOLDER)}")
    logger.info(f"Output folder: {os.path.abspath(OUTPUT_FOLDER)}")
    logger.info(f"Log folder: {os.path.abspath(LOG_FOLDER)}")
    logger.info(f"Session log folder: {os.path.abspath(SESSIONS_LOG_FOLDER)}")
    
    # Start the server
    app.run(debug=True, host="0.0.0.0", port=5000)