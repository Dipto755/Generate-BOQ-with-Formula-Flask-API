import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import re

def extract_boq_formulas():
    """Extract formulas from BOQ sheet and create generalized mapping"""
    
    # File paths
    project_root = Path(__file__).parent
    boq_file_path = project_root / 'data' / 'BOQ3.xlsx'
    output_json_path = project_root / 'boq_formula_mapping.json'
    
    print(f"Reading BOQ file: {boq_file_path}")
    
    if not boq_file_path.exists():
        print(f"ERROR: BOQ file not found at {boq_file_path}")
        return
    
    try:
        # First read with pandas to find last row
        boq_df = pd.read_excel(boq_file_path, sheet_name='BOQ')
        print(f"Loaded BOQ sheet with {len(boq_df)} rows")
        
        # Find last row with data in column A
        last_row = boq_df[boq_df.iloc[:, 0].notna()].index[-1]  # Column A (index 0)
        print(f"Last row with data in column A: {last_row + 1}")
        
        # Use openpyxl to read actual formulas
        wb = load_workbook(boq_file_path, data_only=False)
        boq_ws = wb['BOQ']

        # Extract formulas starting from row 3 (index 2)
        formula_mapping = {}

        for row_idx in range(3, last_row + 2):  # Excel rows are 1-indexed, start from row 3
            item_code_cell = boq_ws.cell(row=row_idx, column=1)  # Column A
            item_code = item_code_cell.value
            
            if item_code is not None:
                item_code_str = str(item_code).strip()
                
                # Get formulas from columns E and F
                formula_E_cell = boq_ws.cell(row=row_idx, column=5)  # Column E
                formula_F_cell = boq_ws.cell(row=row_idx, column=6)  # Column F
                
                # Extract formulas or set to None if empty
                formula_E = formula_E_cell.value if formula_E_cell.value is not None else None
                formula_F = formula_F_cell.value if formula_F_cell.value is not None else None
                
                def process_formula(formula):
                    """Process a formula to remove ALL file references and keep only sheet names"""
                    if not formula or not isinstance(formula, str) or '!' not in formula:
                        return formula
                    
                    # Remove ALL file references [7] from the entire formula
                    # First remove the [7] filename references
                    processed_formula = formula.replace('[7]', '')
                    
                    # Then handle any remaining external references with different patterns
                    # Pattern to match [any_file]sheet!cell_reference and extract just sheet!cell_reference
                    pattern = r'\[([^\]]+)\]([^!]+)!([^+\-*/()=]+)'
                    
                    def remove_filename(match):
                        file_part = match.group(1)  # This part is removed
                        sheet_part = match.group(2)  # Keep the sheet name
                        cell_part = match.group(3)   # Keep the cell reference
                        return f"{sheet_part}!{cell_part}"
                    
                    # Apply the replacement to ALL occurrences throughout the formula
                    processed_formula = re.sub(pattern, remove_filename, processed_formula)
                    
                    return processed_formula
                
                generalized_E = process_formula(formula_E)
                generalized_F = process_formula(formula_F)
                
                # After processing formulas, check if both are None/null
                if generalized_E is None and generalized_F is None:
                    continue  # Skip this item if both formulas are null
                
                formula_mapping[item_code_str] = {
                    'column_E': generalized_E,
                    'column_F': generalized_F,
                    'excel_row': row_idx
                }
        
        print(f"Extracted formulas for {len(formula_mapping)} items")
        
        # Save to JSON
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(formula_mapping, f, indent=2, ensure_ascii=False)
        
        print(f"Formula mapping saved to: {output_json_path}")
        
        # Show sample
        sample_items = list(formula_mapping.keys())[:5]
        print("\nSample extracted formulas:")
        for item in sample_items:
            print(f"  {item}:")
            print(f"    E: {formula_mapping[item]['column_E']}")
            print(f"    F: {formula_mapping[item]['column_F']}")
        
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        print(f"TRACEBACK: {traceback.format_exc()}")

if __name__ == "__main__":
    extract_boq_formulas()