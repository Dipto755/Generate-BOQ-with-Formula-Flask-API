import json
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

def generalize_formula(formula, source_row):
    """
    Replace specific row numbers with {row} placeholder.
    
    Args:
        formula: The Excel formula string
        source_row: The row number to replace (e.g., 15)
    
    Returns:
        Generalized formula string with {row} placeholder
    """
    if not formula:
        return None
    
    # Pattern to match cell references like AC15, AD15, etc.
    # Matches: column letters followed by row number
    pattern = r'([A-Z]+)' + str(source_row) + r'\b'
    
    # Replace with column letter + {row}
    generalized = re.sub(pattern, r'\1{row}', formula)
    
    return generalized

def extract_formulas(excel_path, sheet_name, row_number, start_col, end_col):
    """
    Extract formulas from specified column range and row.
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Name of the sheet to extract from
        row_number: Row number to extract formulas from
        start_col: Starting column (e.g., 'BY')
        end_col: Ending column (e.g., 'KX')
    
    Returns:
        Dictionary mapping column names to generalized formulas
    """
    # Load workbook (data_only=False to access formulas)
    wb = load_workbook(excel_path, data_only=False)
    
    # Get the sheet
    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
    
    # Convert column letters to indices
    start_idx = column_index_from_string(start_col)
    end_idx = column_index_from_string(end_col)
    
    formulas = {}
    
    # Extract formulas from each column
    for col_idx in range(start_idx, end_idx + 1):
        col_letter = get_column_letter(col_idx)
        cell = sheet[f'{col_letter}{row_number}']
        
        # Get the formula (will be None if cell has no formula)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formula = cell.value
            generalized_formula = generalize_formula(formula, row_number)
            formulas[col_letter] = generalized_formula
        elif cell.value:
            # If it has a value but no formula, store as-is (might be a constant)
            formulas[col_letter] = str(cell.value)
    
    wb.close()
    return formulas

def main():
    # Configuration
    excel_file = 'data/Main Carriageway.xlsx'
    sheet_name = 'Quantity'
    source_row = 15
    start_column = 'BY'
    end_column = 'KX'
    output_file = 'formula_template.json'
    
    print(f"Extracting formulas from {excel_file}...")
    print(f"Columns: {start_column} to {end_column}, Row: {source_row}")
    
    # Extract formulas
    formulas = extract_formulas(excel_file, sheet_name, source_row, start_column, end_column)
    
    print(f"Extracted {len(formulas)} formulas")
    
    # Create template structure
    template = {
        "template_name": "main_carriageway_template",
        "source_file": excel_file,
        "sheet_name": sheet_name or "active_sheet",
        "source_row": source_row,
        "column_range": f"{start_column}:{end_column}",
        "formulas": formulas
    }
    
    # Save to JSON
    with open(output_file, 'w') as f:
        json.dump(template, f, indent=2)
    
    print(f"✓ Template saved to {output_file}")
    
    # Show sample formulas
    print("\nSample formulas (first 3):")
    for i, (col, formula) in enumerate(list(formulas.items())[:3]):
        print(f"  {col}: {formula}")

if __name__ == "__main__":
    main()