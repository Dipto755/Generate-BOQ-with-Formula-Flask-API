import json
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

def generalize_sum_formula(formula, source_row):
    """
    Generalize SUM formulas to keep starting row (7) fixed and make ending row dynamic.
    
    Args:
        formula: The Excel formula string (e.g., '=SUM(GY7:GY5090)')
        source_row: The row number to replace (e.g., 15)
    
    Returns:
        Generalized formula string with {row} placeholder for dynamic ending row
    """
    if not formula:
        return None
    
    # Handle SUM formulas with ranges like =SUM(GY7:GY5090)
    sum_pattern = r'=SUM\(([A-Z]+)7:([A-Z]+)(\d+)\)'
    sum_match = re.search(sum_pattern, formula)
    
    if sum_match:
        start_col = sum_match.group(1)
        end_col = sum_match.group(2)
        # Keep starting row as 7, make ending row dynamic
        generalized = f'=SUM({start_col}7:{end_col}{{row}})'
        return generalized
    
    # Handle regular cell references (non-SUM formulas)
    # Pattern to match cell references like AC15, AD15, etc.
    pattern = r'([A-Z]+)' + str(source_row) + r'\b'
    
    # Replace with column letter + {row}
    generalized = re.sub(pattern, r'\1{row}', formula)
    
    return generalized

def extract_formulas_from_columns(excel_path, sheet_name, row_number, columns):
    """
    Extract formulas from specified columns and row.
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Name of the sheet to extract from
        row_number: Row number to extract formulas from
        columns: List of column letters (e.g., ['C', 'GE', 'GF', ..., 'LB'])
    
    Returns:
        Dictionary mapping column names to generalized formulas
    """
    # Load workbook (data_only=False to access formulas)
    wb = load_workbook(excel_path, data_only=False)
    
    # Get the sheet
    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
    
    formulas = {}
    
    # Extract formulas from each specified column
    for col_letter in columns:
        cell = sheet[f'{col_letter}{row_number}']
        
        # Get the formula (will be None if cell has no formula)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formula = cell.value
            generalized_formula = generalize_sum_formula(formula, row_number)
            formulas[col_letter] = generalized_formula
        elif cell.value:
            # If it has a value but no formula, store as-is (might be a constant)
            formulas[col_letter] = str(cell.value)
    
    wb.close()
    return formulas

def generate_column_range(start_col, end_col):
    """
    Generate a list of column letters from start_col to end_col.
    
    Args:
        start_col: Starting column (e.g., 'GE')
        end_col: Ending column (e.g., 'LB')
    
    Returns:
        List of column letters
    """
    start_idx = column_index_from_string(start_col)
    end_idx = column_index_from_string(end_col)
    
    columns = []
    for col_idx in range(start_idx, end_idx + 1):
        columns.append(get_column_letter(col_idx))
    
    return columns

def main():
    # Configuration
    excel_file = 'data/Main Carriageway.xlsx'
    sheet_name = 'Quantity'
    source_row = 5091  # Extract from row 5091 as specified
    single_columns = ['C']  # Single column C
    range_start = 'GE'      # Start of range GE to LB
    range_end = 'LB'        # End of range GE to LB
    output_file = 'formula_final_sum_template.json'
    
    print(f"Extracting SUM formulas from {excel_file}...")
    print(f"Row: {source_row}")
    print(f"Columns: C and {range_start} to {range_end}")
    
    # Generate column list for GE to LB range
    range_columns = generate_column_range(range_start, range_end)
    
    # Combine single columns and range columns
    all_columns = single_columns + range_columns
    
    print(f"Total columns to process: {len(all_columns)}")
    
    # Extract formulas
    formulas = extract_formulas_from_columns(excel_file, sheet_name, source_row, all_columns)
    
    print(f"Extracted {len(formulas)} formulas")
    
    # Create template structure
    template = {
        "template_name": "final_sum_template",
        "source_file": excel_file,
        "sheet_name": sheet_name or "active_sheet",
        "source_row": source_row,
        "columns": {
            "single_columns": single_columns,
            "range": f"{range_start}:{range_end}",
            "all_columns": all_columns
        },
        "formulas": formulas,
        "description": "SUM formulas with fixed start row (7) and dynamic end row"
    }
    
    # Save to JSON
    with open(output_file, 'w') as f:
        json.dump(template, f, indent=2)
    
    print(f"✓ Template saved to {output_file}")
    
    # Show sample formulas
    print("\nSample formulas (first 5):")
    for i, (col, formula) in enumerate(list(formulas.items())[:5]):
        print(f"  {col}: {formula}")
    
    # Show statistics
    sum_formulas = sum(1 for f in formulas.values() if f and 'SUM(' in f)
    print(f"\nStatistics:")
    print(f"  Total formulas: {len(formulas)}")
    print(f"  SUM formulas: {sum_formulas}")
    print(f"  Other formulas: {len(formulas) - sum_formulas}")

if __name__ == "__main__":
    main()
