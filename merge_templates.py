import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, PatternFill, GradientFill, Alignment, Protection, Side
from copy import copy

def merge_templates():
    """Merge main_carriageway and BOQ templates into one file while preserving formatting"""
    
    # File paths
    project_root = Path(__file__).parent
    template_dir = project_root / 'template'
    
    main_template_path = template_dir / 'main_carriageway.xlsx'
    boq_template_path = template_dir / 'BOQ.xlsx'
    output_path = template_dir / 'main_carriageway_and_boq.xlsx'
    
    print(f"Merging templates:")
    print(f"  Main template: {main_template_path}")
    print(f"  BOQ template: {boq_template_path}")
    print(f"  Output: {output_path}")
    
    if not main_template_path.exists():
        print(f"ERROR: Main template not found at {main_template_path}")
        return
    
    if not boq_template_path.exists():
        print(f"ERROR: BOQ template not found at {boq_template_path}")
        return
    
    try:
        # Start with main carriageway template as base
        shutil.copy2(main_template_path, output_path)
        print("Copied main carriageway template as base")
        
        # Load both workbooks
        main_wb = load_workbook(main_template_path)
        boq_wb = load_workbook(boq_template_path)
        output_wb = load_workbook(output_path)
        
        # Copy all sheets from BOQ template to output
        for sheet_name in boq_wb.sheetnames:
            # Skip if sheet already exists (from main template)
            if sheet_name in output_wb.sheetnames:
                print(f"  Skipping '{sheet_name}' - already exists in main template")
                continue
            
            # Get the source sheet from BOQ template
            source_sheet = boq_wb[sheet_name]
            
            # Create new sheet in output workbook
            new_sheet = output_wb.create_sheet(sheet_name)
            
            # Copy all cells with formatting
            print(f"  Copying sheet: {sheet_name}")
            copy_sheet_with_formatting(source_sheet, new_sheet)
        
        # Save the merged workbook
        output_wb.save(output_path)
        
        # Close all workbooks
        main_wb.close()
        boq_wb.close()
        output_wb.close()
        
        print(f"Successfully merged templates to: {output_path}")
        
        # Verify the merge
        verify_wb = load_workbook(output_path)
        print(f"Total sheets in merged file: {len(verify_wb.sheetnames)}")
        print(f"Sheets: {', '.join(verify_wb.sheetnames)}")
        verify_wb.close()
        
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        print(f"TRACEBACK: {traceback.format_exc()}")

def copy_sheet_with_formatting(source_sheet, target_sheet):
    """Copy a sheet with all formatting preserved"""
    
    # Copy column dimensions
    for col_idx, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_idx].width = col_dim.width
        if col_dim.hidden:
            target_sheet.column_dimensions[col_idx].hidden = True
    
    # Copy row dimensions
    for row_idx, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_idx].height = row_dim.height
        if row_dim.hidden:
            target_sheet.row_dimensions[row_idx].hidden = True
    
    # Copy merged cells
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))
    
    # Copy all cells with data and formatting
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet.cell(
                row=cell.row, 
                column=cell.column,
                value=cell.value
            )
            
            # Copy cell formatting
            copy_cell_formatting(cell, target_cell)
    
    # Copy page setup
    try:
        target_sheet.page_setup.orientation = source_sheet.page_setup.orientation
        target_sheet.page_setup.paperSize = source_sheet.page_setup.paperSize
        target_sheet.page_setup.fitToHeight = source_sheet.page_setup.fitToHeight
        target_sheet.page_setup.fitToWidth = source_sheet.page_setup.fitToWidth
    except:
        pass  # Skip if page setup copying fails

def copy_cell_formatting(source_cell, target_cell):
    """Copy formatting from source cell to target cell"""
    
    # Copy font
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=copy(source_cell.font.color)
        )
    
    # Copy border
    if source_cell.border:
        target_cell.border = Border(
            left=copy_border_side(source_cell.border.left),
            right=copy_border_side(source_cell.border.right),
            top=copy_border_side(source_cell.border.top),
            bottom=copy_border_side(source_cell.border.bottom),
            diagonal=copy_border_side(source_cell.border.diagonal),
            diagonal_direction=source_cell.border.diagonal_direction,
            vertical=copy_border_side(source_cell.border.vertical),
            horizontal=copy_border_side(source_cell.border.horizontal)
        )
    
    # Copy fill
    if source_cell.fill:
        if source_cell.fill.patternType is None and source_cell.fill.fill_type is None:
            # No fill
            pass
        elif source_cell.fill.fill_type == 'gradient':
            # Gradient fill
            target_cell.fill = GradientFill(
                type=source_cell.fill.type,
                degree=source_cell.fill.degree,
                left=source_cell.fill.left,
                right=source_cell.fill.right,
                top=source_cell.fill.top,
                bottom=source_cell.fill.bottom,
                stop=[copy(stop) for stop in source_cell.fill.stop]
            )
        else:
            # Pattern fill
            target_cell.fill = PatternFill(
                patternType=source_cell.fill.patternType or source_cell.fill.fill_type,
                fgColor=copy(source_cell.fill.fgColor),
                bgColor=copy(source_cell.fill.bgColor)
            )
    
    # Copy alignment
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )
    
    # Copy number format
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    
    # Copy protection
    if source_cell.protection:
        target_cell.protection = Protection(
            locked=source_cell.protection.locked,
            hidden=source_cell.protection.hidden
        )

def copy_border_side(side):
    """Copy a border side object"""
    if side is None:
        return None
    return Side(
        style=side.style,
        color=copy(side.color)
    )

if __name__ == "__main__":
    merge_templates()