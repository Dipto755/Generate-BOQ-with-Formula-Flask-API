import json
import os
from pathlib import Path
from openpyxl import load_workbook
import sys
import io
from dotenv import load_dotenv

# Add project root to Python path
project_root = os.path.join(os.path.dirname(__file__), '..', '..')
sys.path.append(project_root)

from src.utils.gcs_utils import get_gcs_handler

load_dotenv()
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

class FinalSumApplier:
    """Apply final sum formula templates to output/main_carriageway.xlsx by replacing {row} placeholders."""
    
    def __init__(self, template_path=None, output_excel_path=None):
        if template_path is None:
            # Look for formula_final_sum_template.json in project root
            current_dir = Path(__file__).parent
            template_path = current_dir.parent.parent / "formula_final_sum_template.json"
        
        self.template_path = Path(template_path)
        self.template = self._load_template()
        
        # Initialize session and GCS
        self.session_id = os.getenv('SESSION_ID', 'default')
        self.gcs = get_gcs_handler()
        
        # Determine output filename based on is_merged
        is_merged = os.getenv('IS_MERGED', 'True').lower() == 'true'
        if is_merged:
            self.output_filename = f"{self.session_id}_main_carriageway_and_boq.xlsx"
        else:
            self.output_filename = f"{self.session_id}_main_carriageway.xlsx"
        
        # Handle output path with GCS
        if output_excel_path is None:
            # Download from GCS to temp location
            self.output_gcs_path = self.gcs.get_gcs_path(
                self.session_id, 
                self.output_filename, 
                'output'
            )
            self.output_excel_path = Path(self.gcs.download_to_temp(self.output_gcs_path, suffix='.xlsx'))
        else:
            self.output_excel_path = Path(output_excel_path)
    
    def _load_template(self):
        """Load the formula template from JSON file."""
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        with open(self.template_path, 'r') as f:
            return json.load(f)
    
    def find_last_data_row(self, reference_column='D', start_row=7):
        """
        Find the last row with data in the specified column.
        
        Args:
            reference_column: Column to check for data (default: 'D')
            start_row: Starting row to check from (default: 7)
        
        Returns:
            Last row number with data
        """
        print(f"Looking for output Excel file at: {self.output_excel_path}")
        print(f"Output Excel file exists: {self.output_excel_path.exists()}")
        
        if not self.output_excel_path.exists():
            raise FileNotFoundError(f"Output Excel file not found: {self.output_excel_path}")
        
        # Load output workbook
        output_wb = load_workbook(self.output_excel_path)
        output_sheet_name = "Quantity"
        
        if output_sheet_name not in output_wb.sheetnames:
            raise ValueError(f"Output sheet '{output_sheet_name}' not found in {self.output_excel_path}")
        
        output_sheet = output_wb[output_sheet_name]
        
        # Find last data row
        last_data_row = None
        for row_num in range(start_row, output_sheet.max_row + 1):
            cell_value = output_sheet[f'{reference_column}{row_num}'].value
            if cell_value is not None and str(cell_value).strip():
                last_data_row = row_num
        
        output_wb.close()
        
        if last_data_row is None:
            raise ValueError(f"No data found in column {reference_column} from row {start_row}")
        
        return last_data_row
    
    def apply_formulas_to_last_plus_three(self, reference_column='D', start_row=7):
        """
        Apply formulas to the last row + 3 in the output Excel file.
        
        Args:
            reference_column: Column to check for data (default: 'D')
            start_row: Starting row to check from (default: 7)
        
        Returns:
            Dictionary with statistics about the operation
        """
        # Find the last data row
        last_data_row = self.find_last_data_row(reference_column, start_row)
        # target_row = last_data_row + 3
        target_row = 5091
        
        print(f"Last data row found: {last_data_row}")
        print(f"Target row for formulas: {target_row}")
        
        # Load output workbook
        output_wb = load_workbook(self.output_excel_path)
        output_sheet_name = "Quantity"
        output_sheet = output_wb[output_sheet_name]
        
        formulas = self.template.get("formulas", {})
        
        # Apply formulas to target row
        total_count = 0
        for col_letter, formula_template in formulas.items():
            if formula_template:
                # Replace {row} with the last data row for SUM formulas
                formula = formula_template.replace("{row}", str(last_data_row))
                output_sheet[f"{col_letter}{target_row}"] = formula
                total_count += 1
        
        # Save the output workbook
        output_wb.save(self.output_excel_path)
        output_wb.close()
        
        # Upload to GCS
        self.gcs.upload_file(str(self.output_excel_path), self.output_gcs_path)
        print(f"[GCS] Uploaded to: gs://{self.gcs.bucket.name}/{self.output_gcs_path}")
        
        return {
            "last_data_row": last_data_row,
            "target_row": target_row,
            "formulas_applied": total_count,
            "output_file": str(self.output_excel_path),
            "output_sheet": output_sheet_name
        }
    
    def apply_formulas_with_custom_end_row(self, end_row, target_row_offset=3, start_row=7):
        """
        Apply formulas with custom end row for {row} replacement.
        
        Args:
            end_row: The row number to use for {row} replacement
            target_row_offset: Offset from end_row for target row (default: 3)
            start_row: Starting row to check from (default: 7)
        
        Returns:
            Dictionary with statistics about the operation
        """
        target_row = end_row + target_row_offset
        
        print(f"Using end row: {end_row}")
        print(f"Target row for formulas: {target_row}")
        
        # Load output workbook
        output_wb = load_workbook(self.output_excel_path)
        output_sheet_name = "Quantity"
        output_sheet = output_wb[output_sheet_name]
        
        formulas = self.template.get("formulas", {})
        
        # Apply formulas to target row
        total_count = 0
        for col_letter, formula_template in formulas.items():
            if formula_template:
                # Replace {row} with the specified end row
                formula = formula_template.replace("{row}", str(end_row))
                output_sheet[f"{col_letter}{target_row}"] = formula
                total_count += 1
        
        # Save the output workbook
        output_wb.save(self.output_excel_path)
        output_wb.close()
        
        # Upload to GCS
        self.gcs.upload_file(str(self.output_excel_path), self.output_gcs_path)
        print(f"[GCS] Uploaded to: gs://{self.gcs.bucket.name}/{self.output_gcs_path}")
        
        return {
            "end_row": end_row,
            "target_row": target_row,
            "formulas_applied": total_count,
            "output_file": str(self.output_excel_path),
            "output_sheet": output_sheet_name
        }
    
    def get_template_info(self):
        """Get information about the loaded template."""
        formulas = self.template.get("formulas", {})
        return {
            "template_name": self.template.get("template_name"),
            "source_file": self.template.get("source_file"),
            "sheet_name": self.template.get("sheet_name"),
            "source_row": self.template.get("source_row"),
            "columns": self.template.get("columns"),
            "description": self.template.get("description"),
            "total_formulas": len(formulas)
        }


def main():
    """Command line interface for applying final sum formulas."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Apply final sum formula templates to output/main_carriageway_and_boq.xlsx")
    parser.add_argument("--ref-column", default="D", help="Reference column for auto detection (default: D)")
    parser.add_argument("--start-row", type=int, default=7, help="Starting row to check from (default: 7)")
    parser.add_argument("--template", help="Path to formula_final_sum_template.json")
    parser.add_argument("--output", help="Path to output main_carriageway_and_boq.xlsx file")
    parser.add_argument("--end-row", type=int, help="Custom end row for {row} replacement")
    parser.add_argument("--target-offset", type=int, default=3, help="Offset from end row for target row (default: 3)")
    
    args = parser.parse_args()
    
    applier = FinalSumApplier(template_path=args.template, 
                             output_excel_path=args.output)
    
    info = applier.get_template_info()
    print(f"Loaded template: {info['template_name']}")
    print(f"Total formulas: {info['total_formulas']}")
    print(f"Description: {info['description']}")
    print(f"Column range: {info['columns']['range']}\n")
    
    if args.end_row:
        print(f"Using custom end row: {args.end_row}")
        result = applier.apply_formulas_with_custom_end_row(
            end_row=args.end_row, 
            target_row_offset=args.target_offset,
            start_row=args.start_row
        )
    else:
        print(f"Auto-detecting last data row from column {args.ref_column}...")
        result = applier.apply_formulas_to_last_plus_three(args.ref_column, args.start_row)
    
    print(f"✓ End row for formulas: {result.get('end_row', result.get('last_data_row'))}")
    print(f"✓ Target row: {result['target_row']}")
    print(f"✓ Applied {result['formulas_applied']} formulas")
    print(f"\nFormulas written to: {result['output_file']} (Sheet: {result['output_sheet']})")
    
    return 0


if __name__ == "__main__":
    exit(main())
