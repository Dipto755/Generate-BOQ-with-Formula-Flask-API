import json
from pathlib import Path
from openpyxl import load_workbook
import sys
import io
import os
from dotenv import load_dotenv

# Add project root to Python path
project_root = os.path.join(os.path.dirname(__file__), '..', '..')
sys.path.append(project_root)

from src.utils.gcs_utils import get_gcs_handler

load_dotenv()
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

class FormulaApplier:
    """Apply formula templates to output/main_carriageway.xlsx by replacing {row} placeholders."""
    
    def __init__(self, template_path=None, input_excel_path=None, output_excel_path=None):
        if template_path is None:
            # Look for formula_template.json in project root
            current_dir = Path(__file__).parent
            template_path = current_dir.parent.parent / "formula_template.json"
        
        self.template_path = Path(template_path)
        self.template = self._load_template()
        
        # Initialize session and GCS
        self.session_id = os.getenv('SESSION_ID', 'default')
        self.gcs = get_gcs_handler()
        
        # Determine output filename based on is_merged
        is_merged = os.getenv('IS_MERGED', 'True').lower() == 'true'
        if is_merged:
            self.output_filename = f"{self.session_id}_main_carriageway_and_boq.xlsx"
        else:
            self.output_filename = f"{self.session_id}_main_carriageway.xlsx"
        
        # Handle input/output paths - prefer SESSION_OUTPUT_FILE if available
        if input_excel_path is None:
            # Check for SESSION_OUTPUT_FILE first (local file)
            session_output_file = os.getenv('SESSION_OUTPUT_FILE', '')
            if session_output_file and os.path.exists(session_output_file):
                self.input_excel_path = Path(session_output_file)
                print(f"Using local output file: {session_output_file}")
            else:
                # Fallback: download from GCS (for backward compatibility)
                self.output_gcs_path = self.gcs.get_gcs_path(
                    self.session_id, 
                    self.output_filename, 
                    'output'
                )
                self.input_excel_path = Path(self.gcs.download_to_temp(self.output_gcs_path, suffix='.xlsx'))
                print(f"[GCS] Downloaded output file from GCS: {self.input_excel_path}")
        else:
            self.input_excel_path = Path(input_excel_path)

        if output_excel_path is None:
            # Use same file as input (work on the temp file)
            self.output_excel_path = self.input_excel_path
        else:
            self.output_excel_path = Path(output_excel_path)
        
        # Ensure output directory exists
        self.output_excel_path.parent.mkdir(parents=True, exist_ok=True)
    
    def _load_template(self):
        """Load the formula template from JSON file."""
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        with open(self.template_path, 'r') as f:
            return json.load(f)
    
    def apply_formulas_to_all_data_rows(self, reference_column='D', start_row=7):
        """
        Automatically detect data rows based on a reference column and apply formulas
        to the 'Quantity' sheet starting from specified row.
        
        Args:
            reference_column: Column to check for data (default: 'D')
            start_row: Starting row number for writing formulas (default: 7)
        
        Returns:
            Dictionary with statistics about the operation
        """
        print(f"Looking for input Excel file at: {self.input_excel_path}")
        print(f"Input Excel file exists: {self.input_excel_path.exists()}")
        
        if not self.input_excel_path.exists():
            # Try alternative path - one level up from src/internal to src/data/
            alternative_path = Path(__file__).parent.parent / "data" / self.output_filename
            print(f"Trying alternative path: {alternative_path}")
            print(f"Alternative path exists: {alternative_path.exists()}")
            
            if alternative_path.exists():
                self.input_excel_path = alternative_path
                print(f"Using alternative input path: {self.input_excel_path}")
            else:
                raise FileNotFoundError(f"Input Excel file not found at: {self.input_excel_path}\nAlso tried: {alternative_path}")
        
        # Load input workbook
        input_wb = load_workbook(self.input_excel_path)
        input_sheet_name = "Quantity"
        
        if input_sheet_name not in input_wb.sheetnames:
            raise ValueError(f"Input sheet '{input_sheet_name}' not found in {self.input_excel_path}")
        
        input_sheet = input_wb[input_sheet_name]
        formulas = self.template.get("formulas", {})
        
        # Find data rows from input sheet STARTING FROM start_row
        data_rows = []
        for row_num in range(start_row, input_sheet.max_row + 1):  # Start from start_row instead of 1
            cell_value = input_sheet[f'{reference_column}{row_num}'].value
            if cell_value is not None and str(cell_value).strip():
                data_rows.append(row_num)
        
        if not data_rows:
            raise ValueError(f"No data found in column {reference_column} from row {start_row}")
        
        print(f"Found {len(data_rows)} data rows in input sheet (starting from row {start_row})")
        
        # Load or create output workbook
        if self.output_excel_path.exists():
            output_wb = load_workbook(self.output_excel_path)
            print(f"Loaded existing output file: {self.output_excel_path}")
        else:
            # Create new workbook if output doesn't exist
            output_wb = load_workbook(self.input_excel_path)
            print(f"Created new output file: {self.output_excel_path}")
        
        # Ensure 'Quantity' sheet exists in output
        output_sheet_name = "Quantity"
        if output_sheet_name not in output_wb.sheetnames:
            # Create Quantity sheet if it doesn't exist
            output_wb.create_sheet(output_sheet_name)
            print(f"Created new sheet: {output_sheet_name}")
        
        output_sheet = output_wb[output_sheet_name]
        
        # Apply formulas to output sheet starting from start_row
        total_count = 0
        output_row = start_row
        
        for input_row_num in data_rows:
            for col_letter, formula_template in formulas.items():
                if formula_template:
                    formula = formula_template.replace("{row}", str(input_row_num))
                    output_sheet[f"{col_letter}{output_row}"] = formula
                    total_count += 1
            output_row += 1
        
        # Save the output workbook
        output_wb.save(self.output_excel_path)
        input_wb.close()
        output_wb.close()
        
        # Note: File will be uploaded to GCS at the end of all processing in main.py
        # No need to upload here for efficiency
        
        return {
            "input_first_row": min(data_rows),
            "input_last_row": max(data_rows),
            "input_total_rows": len(data_rows),
            "output_start_row": start_row,
            "output_end_row": output_row - 1,
            "formulas_per_row": len(formulas),
            "total_formulas": total_count,
            "output_file": str(self.output_excel_path),
            "output_sheet": output_sheet_name
        }
    
    def apply_formulas_with_custom_mapping(self, row_mapping, start_row=7):
        """
        Apply formulas with custom row mapping from input to output.
        
        Args:
            row_mapping: Dictionary mapping input_row -> output_row
            start_row: Starting row for output (default: 7)
        
        Returns:
            Dictionary with statistics about the operation
        """
        print(f"Loading input from: {self.input_excel_path}")
        print(f"Writing output to: {self.output_excel_path}")
        
        if not self.input_excel_path.exists():
            raise FileNotFoundError(f"Input Excel file not found: {self.input_excel_path}")
        
        # Load input workbook
        input_wb = load_workbook(self.input_excel_path)
        input_sheet_name = "Quantity"
        
        if input_sheet_name not in input_wb.sheetnames:
            raise ValueError(f"Input sheet '{input_sheet_name}' not found")
        
        input_sheet = input_wb[input_sheet_name]
        formulas = self.template.get("formulas", {})
        
        # Load or create output workbook
        if self.output_excel_path.exists():
            output_wb = load_workbook(self.output_excel_path)
        else:
            output_wb = load_workbook(self.input_excel_path)
        
        # Ensure 'Quantity' sheet exists
        output_sheet_name = "Quantity"
        if output_sheet_name not in output_wb.sheetnames:
            output_wb.create_sheet(output_sheet_name)
        
        output_sheet = output_wb[output_sheet_name]
        
        # Apply formulas using custom mapping
        total_count = 0
        for input_row, output_row in row_mapping.items():
            for col_letter, formula_template in formulas.items():
                if formula_template:
                    formula = formula_template.replace("{row}", str(input_row))
                    output_sheet[f"{col_letter}{output_row}"] = formula
                    total_count += 1
        
        # Save the output workbook
        output_wb.save(self.output_excel_path)
        input_wb.close()
        output_wb.close()
        
        # Note: File will be uploaded to GCS at the end of all processing in main.py
        # No need to upload here for efficiency
        
        return {
            "total_mappings": len(row_mapping),
            "total_formulas": total_count,
            "output_file": str(self.output_excel_path),
            "output_sheet": output_sheet_name
        }
    
    def get_template_info(self):
        """Get information about the loaded template."""
        formulas = self.template.get("formulas", {})
        return {
            "template_name": self.template.get("template_name"),
            "source_file": self.template.get("source_file"),
            "sheet_name": self.template.get("sheet_name"),
            "source_row": self.template.get("source_row"),
            "column_range": self.template.get("column_range"),
            "total_formulas": len(formulas)
        }


def main():
    """Command line interface for applying formulas."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Apply formula templates to output/main_carriageway_and_boq.xlsx")
    parser.add_argument("--ref-column", default="D", help="Reference column for auto detection")
    parser.add_argument("--start-row", type=int, default=7, help="Starting row for output (default: 7)")
    parser.add_argument("--template", help="Path to formula template JSON")
    parser.add_argument("--input", help="Path to input main_carriageway_and_boq.xlsx file")
    parser.add_argument("--output", help="Path to output main_carriageway_and_boq.xlsx file")
    
    args = parser.parse_args()
    
    applier = FormulaApplier(template_path=args.template, 
                           input_excel_path=args.input, 
                           output_excel_path=args.output)
    
    info = applier.get_template_info()
    print(f"Loaded template: {info['template_name']}")
    print(f"Total formulas: {info['total_formulas']}")
    print(f"Column range: {info['column_range']}\n")
    
    print(f"Auto-detecting data rows from column {args.ref_column}...")
    result = applier.apply_formulas_to_all_data_rows(args.ref_column, args.start_row)
    
    print(f"✓ Input data rows: {result['input_first_row']} to {result['input_last_row']}")
    print(f"✓ Output rows: {result['output_start_row']} to {result['output_end_row']}")
    print(f"✓ Applied {result['total_formulas']} formulas to {result['input_total_rows']} rows")
    print(f"\nFormulas written to: {result['output_file']} (Sheet: {result['output_sheet']})")
    
    return 0


if __name__ == "__main__":
    exit(main())
