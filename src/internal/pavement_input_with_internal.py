"""
Geogrid Calculator
Reads Pavement_Input.xlsx to check for Geogrid conditions
Calculates columns KY, KZ, LA, LB in main_carriageway.xlsx based on formulas
"""

import pandas as pd
import os
import sys
import io
from dotenv import load_dotenv

load_dotenv()
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


# ============================================================================
# FILE PATHS - Update these to match your folder structure
# ============================================================================

# NEW CODE:
script_dir = os.path.dirname(os.path.abspath(__file__))
# Use session directories from environment, fallback to original paths
data_dir = os.getenv('SESSION_DATA_DIR', os.path.join(script_dir, '..', '..', 'data'))
output_file = os.getenv('SESSION_OUTPUT_FILE', os.path.join(script_dir, '..', '..', 'output', 'main_carriageway_and_boq.xlsx'))

PAVEMENT_INPUT_FILE = os.path.join(data_dir, 'Pavement Input.xlsx')
MAIN_CARRIAGEWAY_FILE = output_file
OUTPUT_EXCEL = output_file


# ============================================================================
# STEP 1: Check Pavement_Input.xlsx for Geogrid Conditions
# ============================================================================

def check_geogrid_conditions(pavement_input_file):
    """
    Reads Pavement_Input.xlsx and checks if:
    - E9 contains "Geogrid Reinforced GSB"
    - E10 contains "Geogrid Reinforced WMM"
    - B9 contains "Geogrid Reinforced GSB"
    - B10 contains "Geogrid Reinforced WMM"
    Returns: dictionary with boolean flags
    """
    print("="*80)
    print("STEP 1: Checking Geogrid Conditions in Pavement Input")
    print("="*80)
    
    # Read the Excel file without headers
    df = pd.read_excel(pavement_input_file, header=None)
    
    print("[OK] Read Pavement_Input.xlsx:", len(df), "total rows")
    
    # Initialize conditions
    conditions = {
        'e9_geogrid_gsb': False,
        'e10_geogrid_wmm': False,
        'b9_geogrid_gsb': False,
        'b10_geogrid_wmm': False
    }
    
    # Check E9 (row index 8, column 4)
    if len(df) > 8:
        e9_value = df.iloc[8, 4]  # E9
        if pd.notna(e9_value) and "Geogrid Reinforced GSB" in str(e9_value):
            conditions['e9_geogrid_gsb'] = True
            print(f"[OK] E9 contains 'Geogrid Reinforced GSB': {e9_value}")
        else:
            print(f"[OK] E9 value: {e9_value} (not Geogrid Reinforced GSB)")
    
    # Check E10 (row index 9, column 4)
    if len(df) > 9:
        e10_value = df.iloc[9, 4]  # E10
        if pd.notna(e10_value) and "Geogrid Reinforced WMM" in str(e10_value):
            conditions['e10_geogrid_wmm'] = True
            print(f"[OK] E10 contains 'Geogrid Reinforced WMM': {e10_value}")
        else:
            print(f"[OK] E10 value: {e10_value} (not Geogrid Reinforced WMM)")
    
    # Check B9 (row index 8, column 1)
    if len(df) > 8:
        b9_value = df.iloc[8, 1]  # B9
        if pd.notna(b9_value) and "Geogrid Reinforced GSB" in str(b9_value):
            conditions['b9_geogrid_gsb'] = True
            print(f"[OK] B9 contains 'Geogrid Reinforced GSB': {b9_value}")
        else:
            print(f"[OK] B9 value: {b9_value} (not Geogrid Reinforced GSB)")
    
    # Check B10 (row index 9, column 1)
    if len(df) > 9:
        b10_value = df.iloc[9, 1]  # B10
        if pd.notna(b10_value) and "Geogrid Reinforced WMM" in str(b10_value):
            conditions['b10_geogrid_wmm'] = True
            print(f"[OK] B10 contains 'Geogrid Reinforced WMM': {b10_value}")
        else:
            print(f"[OK] B10 value: {b10_value} (not Geogrid Reinforced WMM)")
    
    print("\n[OK] Geogrid Conditions Summary:")
    print(f"  E9 Geogrid GSB: {conditions['e9_geogrid_gsb']}")
    print(f"  E10 Geogrid WMM: {conditions['e10_geogrid_wmm']}")
    print(f"  B9 Geogrid GSB: {conditions['b9_geogrid_gsb']}")
    print(f"  B10 Geogrid WMM: {conditions['b10_geogrid_wmm']}")
    
    return conditions


# ============================================================================
# STEP 2: Calculate Geogrid Columns
# ============================================================================

def find_last_row_with_data(ws, column_letter):
    """
    Find the last row with data in the specified column
    Args:
        ws: Worksheet object
        column_letter: Excel column letter (e.g., 'D')
    Returns:
        Last row number that contains data
    """
    from openpyxl.utils import column_index_from_string
    
    col_idx = column_index_from_string(column_letter)
    last_row = 0
    
    # Iterate from bottom to top to find last non-empty cell
    for row_idx in range(ws.max_row, 0, -1):
        cell_value = ws.cell(row_idx, col_idx).value
        if cell_value is not None and cell_value != '':
            last_row = row_idx
            break
    
    return last_row

def calculate_geogrid_columns(main_carriageway_file, conditions, output_file):
    """
    Reads main_carriageway_and_boq.xlsx and calculates geogrid columns based on conditions
    Updates only columns KY, KZ, LA, LB without touching other columns
    """
    print("\n" + "="*80)
    print("STEP 2: Calculating Geogrid Columns")
    print("="*80)
    
    from openpyxl import load_workbook
    
    # Load workbook
    wb = load_workbook(main_carriageway_file)
    ws = wb['Quantity']
    
    print(f"[OK] Loaded workbook: {main_carriageway_file}")
    print("  Sheet: Quantity")
    print(f"  Max row: {ws.max_row}, Max column: {ws.max_column}")
    
    # Find last row with data using column D as reference
    last_row_with_data = find_last_row_with_data(ws, 'D')
    print(f"[OK] Last row with data in column D: {last_row_with_data}")
    
    if last_row_with_data == 0:
        print("[WARNING] No data found in column D, using max_row instead")
        last_row_with_data = ws.max_row
    
    # Column letters (Excel columns, 1-indexed)
    LENGTH_COL = 3      # Column C
    DL_COL = 116        # Column DL
    DS_COL = 123        # Column DS
    EE_COL = 135        # Column EE
    EJ_COL = 140        # Column EJ
    FD_COL = 160        # Column FD
    FF_COL = 162        # Column FF
    FS_COL = 175        # Column FS
    FY_COL = 181        # Column FY
    
    KY_COL = 311        # Column KY
    KZ_COL = 312        # Column KZ
    LA_COL = 313        # Column LA
    LB_COL = 314        # Column LB
    
    # Data starts from row 7
    start_row = 7
    
    print(f"\n[OK] Calculating geogrid values from row {start_row} to row {last_row_with_data}...")
    
    row_count = 0
    for row_idx in range(start_row, last_row_with_data + 1):
        # Get length value
        length_cell = ws.cell(row_idx, LENGTH_COL)
        length = length_cell.value if length_cell.value is not None else 0
        
        # Skip empty rows (if length column is empty, skip this row)
        if length == 0 or length is None or length == '':
            continue
        
        # Get column values
        dl_val = ws.cell(row_idx, DL_COL).value or 0
        ds_val = ws.cell(row_idx, DS_COL).value or 0
        ee_val = ws.cell(row_idx, EE_COL).value or 0
        ej_val = ws.cell(row_idx, EJ_COL).value or 0
        ff_val = ws.cell(row_idx, FF_COL).value or 0
        fd_val = ws.cell(row_idx, FD_COL).value or 0
        fy_val = ws.cell(row_idx, FY_COL).value or 0
        fs_val = ws.cell(row_idx, FS_COL).value or 0
        
        # Calculate KY
        ky_val = ((dl_val if conditions['e9_geogrid_gsb'] else 0) + 
                  (ds_val if conditions['e10_geogrid_wmm'] else 0)) * length
        ws.cell(row_idx, KY_COL).value = ky_val
        
        # Calculate KZ
        kz_val = ((ee_val if conditions['b9_geogrid_gsb'] else 0) + 
                  (ej_val if conditions['b10_geogrid_wmm'] else 0)) * length
        ws.cell(row_idx, KZ_COL).value = kz_val
        
        # Calculate LA
        la_val = ((ff_val if conditions['b9_geogrid_gsb'] else 0) + 
                  (fd_val if conditions['b10_geogrid_wmm'] else 0)) * length
        ws.cell(row_idx, LA_COL).value = la_val
        
        # Calculate LB
        lb_val = ((fy_val if conditions['e9_geogrid_gsb'] else 0) + 
                  (fs_val if conditions['e10_geogrid_wmm'] else 0)) * length
        ws.cell(row_idx, LB_COL).value = lb_val
        
        row_count += 1
        
        # Progress indicator
        if row_count % 200 == 0:
            print(f"  Processed {row_count} rows...")
    
    print(f"[OK] Geogrid calculations completed for {row_count} rows")
    
    # Save workbook
    print(f"\n[OK] Saving to {output_file}...")
    wb.save(output_file)
    
    print("[OK] Saved!")
    
    return row_count


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main function to execute all steps"""
    print("\n" + "="*80)
    print("GEOGRID CALCULATOR")
    print("="*80)
    print("Configuration:")
    print("  • Pavement_Input.xlsx: Check for Geogrid conditions")
    print("  • main_carriageway_and_boq.xlsx: Calculate columns KY, KZ, LA, LB")
    print("="*80 + "\n")
    
    try:
        # Step 1: Check geogrid conditions
        conditions = check_geogrid_conditions(PAVEMENT_INPUT_FILE)
        
        # Step 2: Calculate geogrid columns
        row_count = calculate_geogrid_columns(MAIN_CARRIAGEWAY_FILE, conditions, OUTPUT_EXCEL)
        
        # Success summary
        print("\n" + "="*80)
        print("SUCCESS! [OK]")
        print("="*80)
        print("Output file:", OUTPUT_EXCEL)
        print("Total rows processed:", row_count)
        print("Columns updated: KY, KZ, LA, LB")
        
    except FileNotFoundError as e:
        print("\n[ERROR] File not found")
        print(" ", e)
        print("\nPlease check:")
        print("  1. Files exist in the data folder")
        print("  2. File names match exactly")
    except Exception as e:
        print("\n[ERROR]:", e)
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
