#!/usr/bin/env python3
"""
Excel Formula Recalculation Script for main_carriageway.xlsx

This script recalculates all formulas in output/main_carriageway.xlsx using LibreOffice Calc
and scans for formula errors in the Quantity sheet.

Usage:
    python recalc.py [timeout_seconds]
"""

import sys
import json
import subprocess
import time
import os
import tempfile
from pathlib import Path
import platform


class ExcelRecalculator:
    """Recalculate Excel formulas using LibreOffice and detect errors."""
    
    # Excel error types to detect
    ERROR_TYPES = ['#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NULL!', '#NUM!']
    
    def __init__(self):
        self.system = platform.system()
        self.libreoffice_path = self._find_libreoffice()
        
        # Set paths for output file
        current_dir = Path(__file__).parent
        self.output_excel_path = current_dir.parent.parent / "output" / "main_carriageway.xlsx"
        
        # Ensure output directory exists
        self.output_excel_path.parent.mkdir(parents=True, exist_ok=True)
    
    def _find_libreoffice(self):
        """Find LibreOffice installation path."""
        if self.system == 'Linux':
            return '/usr/bin/libreoffice'
        elif self.system == 'Darwin':  # macOS
            return '/Applications/LibreOffice.app/Contents/MacOS/soffice'
        elif self.system == 'Windows':
            return 'C:\\Program Files\\LibreOffice\\program\\soffice.exe'
        return None
    
    def _check_libreoffice(self):
        """Check if LibreOffice is available."""
        if not self.libreoffice_path or not os.path.exists(self.libreoffice_path):
            raise RuntimeError(
                "LibreOffice not found. Please install LibreOffice:\n"
                "  Ubuntu/Debian: sudo apt-get install libreoffice\n"
                "  macOS: brew install --cask libreoffice\n"
                "  Or download from: https://www.libreoffice.org/download/"
            )
    
    def recalculate(self, timeout=60):
        """
        Recalculate all formulas in output/main_carriageway.xlsx.
        
        Args:
            timeout: Timeout in seconds (default: 60)
        
        Returns:
            Dictionary with recalculation results
        """
        self._check_libreoffice()
        
        print(f"Looking for output Excel file at: {self.output_excel_path}")
        print(f"Output Excel file exists: {self.output_excel_path.exists()}")
        
        if not self.output_excel_path.exists():
            raise FileNotFoundError(f"Output Excel file not found at: {self.output_excel_path}")
        
        print(f"Recalculating formulas in: {self.output_excel_path}")
        
        try:
            # Convert file using LibreOffice headless mode (this recalculates formulas)
            cmd = [
                self.libreoffice_path,
                '--headless',
                '--calc',
                '--convert-to', 'xlsx',
                '--outdir', str(self.output_excel_path.parent),
                str(self.output_excel_path)
            ]
            
            print("Running LibreOffice...")
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=timeout
            )
            
            if result.returncode != 0:
                print(f"Warning: LibreOffice returned code {result.returncode}")
                if result.stderr:
                    print(f"stderr: {result.stderr}")
            
            print("✓ Formulas recalculated")
            
        except subprocess.TimeoutExpired:
            raise RuntimeError(f"LibreOffice timed out after {timeout} seconds")
        
        except Exception as e:
            raise RuntimeError(f"Error running LibreOffice: {str(e)}")
        
        # Scan for errors in Quantity sheet starting from row 7
        error_report = self.scan_errors()
        
        return error_report
    
    def scan_errors(self, start_row=7):
        """
        Scan Quantity sheet for formula errors starting from specified row.
        
        Args:
            start_row: Starting row to scan from (default: 7)
        
        Returns:
            Dictionary with error report
        """
        from openpyxl import load_workbook
        
        print(f"Scanning for formula errors in Quantity sheet from row {start_row}...")
        
        wb = load_workbook(self.output_excel_path, data_only=True)
        
        error_summary = {}
        total_errors = 0
        total_formulas = 0
        scanned_cells = 0
        
        sheet_name = "Quantity"
        
        if sheet_name not in wb.sheetnames:
            print(f"⚠ Warning: Sheet '{sheet_name}' not found in {self.output_excel_path}")
            wb.close()
            return {
                "status": "sheet_not_found",
                "total_errors": 0,
                "total_formulas": 0,
                "scanned_cells": 0,
                "file": str(self.output_excel_path),
                "sheet": sheet_name,
                "start_row": start_row
            }
        
        sheet = wb[sheet_name]
        
        # Scan only from start_row onwards
        for row_num, row in enumerate(sheet.iter_rows(), 1):
            if row_num < start_row:
                continue
                
            for cell in row:
                scanned_cells += 1
                if cell.value is not None:
                    cell_value = str(cell.value)
                    
                    # Check if it's an error
                    for error_type in self.ERROR_TYPES:
                        if error_type in cell_value:
                            total_errors += 1
                            
                            if error_type not in error_summary:
                                error_summary[error_type] = {
                                    'count': 0,
                                    'locations': []
                                }
                            
                            error_summary[error_type]['count'] += 1
                            error_summary[error_type]['locations'].append(
                                f"{sheet_name}!{cell.coordinate}"
                            )
        
        # Count formulas in Quantity sheet from start_row
        wb_formulas = load_workbook(self.output_excel_path, data_only=False)
        
        if sheet_name in wb_formulas.sheetnames:
            sheet_formulas = wb_formulas[sheet_name]
            
            for row_num, row in enumerate(sheet_formulas.iter_rows(), 1):
                if row_num < start_row:
                    continue
                    
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        total_formulas += 1
        
        wb.close()
        wb_formulas.close()
        
        # Build report
        if total_errors > 0:
            status = "errors_found"
            print(f"⚠ Found {total_errors} formula errors in Quantity sheet (rows {start_row}+)")
        else:
            status = "success"
            print(f"✓ No formula errors found in Quantity sheet (rows {start_row}+)")
        
        report = {
            "status": status,
            "total_errors": total_errors,
            "total_formulas": total_formulas,
            "scanned_cells": scanned_cells,
            "file": str(self.output_excel_path),
            "sheet": sheet_name,
            "start_row": start_row,
            "scanned_range": f"Row {start_row} to end"
        }
        
        if total_errors > 0:
            report["error_summary"] = error_summary
        
        return report
    
    def scan_specific_range(self, sheet_name="Quantity", start_row=7, end_row=None, columns=None):
        """
        Scan a specific range in the output file for errors.
        
        Args:
            sheet_name: Sheet to scan (default: "Quantity")
            start_row: Starting row (default: 7)
            end_row: Ending row (optional)
            columns: Specific columns to scan (optional)
        
        Returns:
            Dictionary with error report
        """
        from openpyxl import load_workbook
        
        print(f"Scanning {sheet_name} sheet from row {start_row}...")
        
        wb = load_workbook(self.output_excel_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {
                "status": "sheet_not_found",
                "total_errors": 0,
                "file": str(self.output_excel_path),
                "sheet": sheet_name
            }
        
        sheet = wb[sheet_name]
        error_summary = {}
        total_errors = 0
        scanned_cells = 0
        
        # Determine range to scan
        max_row = end_row if end_row else sheet.max_row
        
        for row_num in range(start_row, max_row + 1):
            if columns:
                # Scan specific columns
                for col in columns:
                    cell = sheet[f"{col}{row_num}"]
                    scanned_cells += 1
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        for error_type in self.ERROR_TYPES:
                            if error_type in cell_value:
                                total_errors += 1
                                if error_type not in error_summary:
                                    error_summary[error_type] = {'count': 0, 'locations': []}
                                error_summary[error_type]['count'] += 1
                                error_summary[error_type]['locations'].append(f"{sheet_name}!{cell.coordinate}")
            else:
                # Scan entire row
                for cell in sheet[row_num]:
                    scanned_cells += 1
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        for error_type in self.ERROR_TYPES:
                            if error_type in cell_value:
                                total_errors += 1
                                if error_type not in error_summary:
                                    error_summary[error_type] = {'count': 0, 'locations': []}
                                error_summary[error_type]['count'] += 1
                                error_summary[error_type]['locations'].append(f"{sheet_name}!{cell.coordinate}")
        
        wb.close()
        
        report = {
            "status": "errors_found" if total_errors > 0 else "success",
            "total_errors": total_errors,
            "scanned_cells": scanned_cells,
            "file": str(self.output_excel_path),
            "sheet": sheet_name,
            "range": f"Rows {start_row}-{max_row}" + (f", Columns {columns}" if columns else "")
        }
        
        if total_errors > 0:
            report["error_summary"] = error_summary
        
        return report


def main():
    """Main entry point."""
    timeout = int(sys.argv[1]) if len(sys.argv) > 1 else 60
    
    try:
        recalculator = ExcelRecalculator()
        
        # Recalculate formulas
        result = recalculator.recalculate(timeout=timeout)
        
        # Print JSON report
        print("\nRecalculation Report:")
        print(json.dumps(result, indent=2))
        
        # Return appropriate exit code
        return 0 if result['status'] == 'success' else 1
    
    except Exception as e:
        print(f"Error: {str(e)}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())