import os
import openpyxl
from openpyxl.utils import get_column_letter

def clean_excel_file():
    print("=" * 50)
    print("Starting Excel file cleaning process...")
    print("=" * 50)
    
    # Define paths
    current_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(current_dir)
    data_dir = os.path.join(root_dir, 'data')
    template_dir = os.path.join(root_dir, 'template')
    
    print(f"Current directory: {current_dir}")
    print(f"Root directory: {root_dir}")
    print(f"Data directory: {data_dir}")
    print(f"Template directory: {template_dir}")
    
    # Input and output file paths
    input_file = os.path.join(data_dir, 'Main Carriageway.xlsx')
    output_file = os.path.join(template_dir, 'main_carriageway.xlsx')
    
    print(f"\nInput file path: {input_file}")
    print(f"Output file path: {output_file}")
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"❌ Error: Input file not found at {input_file}")
        return
    else:
        print("✅ Input file found")
    
    # Create template directory if it doesn't exist
    os.makedirs(template_dir, exist_ok=True)
    print(f"✅ Template directory ready at {template_dir}")
    
    # Load the workbook
    try:
        print("\n📂 Loading workbook...")
        wb = openpyxl.load_workbook(input_file)
        print(f"✅ Workbook loaded successfully")
        print(f"📊 Sheets found: {wb.sheetnames}")
    except Exception as e:
        print(f"❌ Error loading workbook: {e}")
        return
    
    # Check if 'Quantity' sheet exists
    if 'Quantity' not in wb.sheetnames:
        print(f"❌ Error: Sheet 'Quantity' not found in workbook")
        print(f"Available sheets: {wb.sheetnames}")
        return
    else:
        print("✅ Sheet 'Quantity' found")
    
    # Process only the 'Quantity' sheet
    sheet_name = 'Quantity'
    print(f"\n{'='*20} Processing sheet: {sheet_name} {'='*20}")
    sheet = wb[sheet_name]
    print(f"Sheet max row: {sheet.max_row}")
    print(f"Sheet max column: {sheet.max_column}")
    
    # First, unmerge all merged cells that start at row 7 or below
    print("\n🔍 Checking for merged cells...")
    merged_ranges = list(sheet.merged_cells.ranges)
    print(f"Total merged ranges found: {len(merged_ranges)}")
    
    unmerged_count = 0
    for merged_range in merged_ranges:
        if merged_range.min_row >= 7:
            print(f"  Unmerging: {merged_range} (starts at row {merged_range.min_row})")
            sheet.unmerge_cells(str(merged_range))
            unmerged_count += 1
    
    print(f"✅ Unmerged {unmerged_count} ranges that start at row 7 or below")
    
    # Now clear all cells from row 7 onwards
    print(f"\n🧹 Clearing cells from row 7 to {sheet.max_row}...")
    cells_cleared = 0
    for row_idx, row in enumerate(sheet.iter_rows(min_row=7, max_row=sheet.max_row), start=7):
        for col_idx, cell in enumerate(row, start=1):
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                if cell.value is not None:
                    print(f"  Clearing cell {get_column_letter(col_idx)}{row_idx}: '{cell.value}'")
                    cell.value = None
                    cells_cleared += 1
    
    print(f"✅ Cleared {cells_cleared} non-merged cells")
    
    # Also clear any remaining merged cells that might have been unmerged
    # print("\n🔍 Final check for any remaining values...")
    # remaining_values = 0
    # for row in range(7, sheet.max_row + 1):
    #     for col in range(1, sheet.max_column + 1):
    #         cell = sheet.cell(row=row, column=col)
    #         if cell.value is not None:
    #             print(f"  Clearing remaining value at {get_column_letter(col)}{row}: '{cell.value}'")
    #             cell.value = None
    #             remaining_values += 1
    
    # print(f"✅ Cleared {remaining_values} remaining values")
    # print(f"✅ Sheet '{sheet_name}' processing complete")
    
    # Save the cleaned workbook
    try:
        print(f"\n💾 Saving cleaned workbook to {output_file}...")
        wb.save(output_file)
        print("✅ File saved successfully!")
        print("=" * 50)
        print("🎉 Cleaning process completed successfully!")
        print("=" * 50)
    except Exception as e:
        print(f"❌ Error saving workbook: {e}")
        print("=" * 50)

if __name__ == "__main__":
    clean_excel_file()