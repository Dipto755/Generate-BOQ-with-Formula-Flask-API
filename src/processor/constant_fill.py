"""
Constant Fill Processor
Fills constant values in specified columns of main_carriageway.xlsx
Applies predefined constants to various cells based on requirements
"""

import pandas as pd
import os
import sys
import io
import time
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def wait_for_file_ready(file_path, max_wait_seconds=15):
    """Wait for file to be ready for access"""
    import os
    start_time = time.time()
    
    print(f"[INFO] Waiting for file to be ready: {os.path.basename(file_path)}")
    
    while time.time() - start_time < max_wait_seconds:
        try:
            # Try to open file in read mode to check if it's ready
            with open(file_path, 'rb') as f:
                pass
            # Additional check: try to read first few bytes
            with open(file_path, 'rb') as f:
                f.read(1024)  # Try to read first 1KB
            print(f"[OK] File is ready for access: {os.path.basename(file_path)}")
            return True
        except (IOError, PermissionError) as e:
            print(f"[WAIT] File not ready, waiting... ({time.time() - start_time:.1f}s elapsed)")
            time.sleep(1.0)
        except Exception as e:
            print(f"[ERROR] Unexpected error waiting for file: {e}")
            return False
    
    print(f"[ERROR] Timeout waiting for file to be ready: {file_path}")
    return False

def safe_workbook_operation(file_path, operation_func, max_retries=5):
    """Safely perform workbook operations with enhanced retry logic"""
    import os
    from openpyxl import load_workbook
    
    for attempt in range(max_retries):
        try:
            # Wait for file to be ready
            if not wait_for_file_ready(file_path, max_wait_seconds=10):
                raise IOError(f"File not ready for access: {file_path}")
            
            print(f"[INFO] Attempting workbook operation (attempt {attempt + 1}/{max_retries})")
            
            wb = load_workbook(file_path)
            result = operation_func(wb)
            
            # Save and close with additional delay
            wb.save(file_path)
            wb.close()
            
            # Wait longer after save to ensure file system completion
            time.sleep(2.0)
            
            print(f"[OK] Workbook operation completed successfully")
            return result
            
        except Exception as e:
            print(f"[WARNING] Attempt {attempt + 1} failed: {e}")
            if attempt == max_retries - 1:
                print(f"[ERROR] All attempts failed for {os.path.basename(file_path)}")
                raise e
            
            # Exponential backoff with jitter
            wait_time = (2 ** attempt) * 0.5 + (attempt * 0.1)
            print(f"[INFO] Waiting {wait_time:.1f}s before retry...")
            time.sleep(wait_time)

def main(session_id=None):
    """Main function to execute constant fill processing"""
    import sys
    sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))
    from api.session_manager import session_manager
    
    # Get session information
    if session_id:
        session = session_manager.get_session(session_id)
        if not session:
            print(f"[ERROR] Session {session_id} not found")
            return
        session_dir = session["output_dir"]
    else:
        # Fallback to original paths if no session_id
        session_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'output')
    
    # Main carriageway file in session directory with session_id suffix
    main_carriageway_file = os.path.join(session_dir, f'main_carriageway_{session_id}.xlsx')
    output_file = os.path.join(session_dir, f'main_carriageway_{session_id}.xlsx')
    
    try:
        # Step 1: Apply constant fills
        apply_constant_fills(main_carriageway_file, output_file)
        
        print("\n" + "="*80)
        print("SUCCESS! [OK]")
        print("="*80)
        print("Output file:", output_file)
        print("Constant values applied successfully")
        
    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()

def apply_constant_fills(main_carriageway_file, output_file):
    """
    Apply constant values to specified columns in main_carriageway.xlsx
    """
    print("="*80)
    print("CONSTANT FILL PROCESSOR")
    print("="*80)
    
    from openpyxl import load_workbook
    
    # Define constant values and their target columns
    # These are example constants - adjust based on actual requirements
    CONSTANT_VALUES = {
        # Example constant fills (adjust column numbers and values as needed)
        57: 1.0,    # Column BF - Example constant
        58: 2.5,    # Column BG - Example constant
        59: 0.75,   # Column BH - Example constant
        60: 1.2,    # Column BI - Example constant
        # Add more constants as needed
    }
    
    print("\nSTEP 1: Loading Main Carriageway Workbook")
    print("-" * 40)
    
    def process_workbook(wb):
        """Inner function to process the workbook"""
        from openpyxl import load_workbook
        
        ws = wb['Quantity']
        
        print(f"[OK] Loaded workbook: {main_carriageway_file}")
        print(f"  Sheet: Quantity")
        print(f"  Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        print(f"\nSTEP 2: Applying Constant Values")
        print("-" * 40)
        print(f"[OK] Applying {len(CONSTANT_VALUES)} constant values")
        
        # Define data range
        LENGTH_COL = 3   # Column C
        start_row = 7
        
        # Find last row with data
        last_row = 0
        for row_idx in range(start_row, ws.max_row + 1):
            cell_value = ws.cell(row_idx, LENGTH_COL).value
            if cell_value is not None and cell_value != '':
                last_row = row_idx
        
        if last_row == 0:
            print("[WARNING] No data found in column C")
            return 0
        
        print(f"[OK] Processing rows {start_row} to {last_row}")
        
        # Apply constant values
        applied_count = 0
        for col_idx, constant_value in CONSTANT_VALUES.items():
            # Apply constant to all data rows in this column
            for row_idx in range(start_row, last_row + 1):
                # Check if this row has data (using length column as reference)
                length_value = ws.cell(row_idx, LENGTH_COL).value
                if length_value is not None and length_value != '' and length_value != 0:
                    ws.cell(row_idx, col_idx).value = constant_value
                    applied_count += 1
            
            # Get column letter for display
            col_letter = get_excel_column_letter(col_idx)
            print(f"  Applied constant {constant_value} to column {col_letter}")
        
        print(f"[OK] Applied constants to {applied_count} cells")
        
        # Additional: Apply specific cell constants if needed
        apply_specific_cell_constants(ws)
        
        return applied_count
    
    # Use safe workbook operation
    try:
        return safe_workbook_operation(main_carriageway_file, process_workbook)
    except Exception as e:
        print(f"[ERROR] Failed to apply constant fills: {e}")
        raise

def apply_specific_cell_constants(ws):
    """
    Apply constants to specific cells (row, column) if needed
    This is for cells that need specific values regardless of data rows
    """
    # Example specific cell constants (adjust as needed)
    SPECIFIC_CONSTANTS = {
        # (row, column): value
        (5091, 70): "TOTAL",  # Example: Total label in specific cell
        (5091, 71): 1000,     # Example: multiplier value
        # Add more specific constants as needed
    }
    
    for (row_idx, col_idx), value in SPECIFIC_CONSTANTS.items():
        try:
            ws.cell(row_idx, col_idx).value = value
            col_letter = get_excel_column_letter(col_idx)
            print(f"  Applied specific constant {value} to cell {col_letter}{row_idx}")
        except Exception as e:
            print(f"  [WARNING] Could not apply constant to cell ({row_idx}, {col_idx}): {e}")

def get_excel_column_letter(col_idx):
    """
    Convert 1-based column index to Excel column letter
    """
    if col_idx < 1:
        return ""
    
    col_letter = ""
    while col_idx > 0:
        col_idx -= 1
        col_letter = chr(65 + (col_idx % 26)) + col_letter
        col_idx //= 26
    
    return col_letter

def load_constants_from_config(config_file=None):
    """
    Load constant values from a configuration file if available
    This allows for easy modification without changing code
    """
    import json
    
    if not config_file:
        # Look for config file in same directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        config_file = os.path.join(script_dir, 'constant_fill_config.json')
    
    if not os.path.exists(config_file):
        print(f"[INFO] No config file found at {config_file}, using default constants")
        return {}
    
    try:
        with open(config_file, 'r') as f:
            config = json.load(f)
        
        print(f"[OK] Loaded constants from config file: {config_file}")
        return config.get('constants', {})
        
    except Exception as e:
        print(f"[WARNING] Could not load config file: {e}")
        return {}

if __name__ == "__main__":
    main()
