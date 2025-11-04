"""
Embankment Height Calculator
Reads EMB_HEIGHT.xlsx to calculate embankment heights for each TCS chainage
Calculates columns BJ, BK, BL, BM in main_carriageway.xlsx based on level differences
"""

import pandas as pd
import os
import sys
import io
import time

def main(session_id=None):
    """Main function to execute embankment height processing"""
    import sys
    sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))
    from api.session_manager import session_manager
    
    # Get session information
    if session_id:
        session = session_manager.get_session(session_id)
        if not session:
            print(f"[ERROR] Session {session_id} not found")
            return
        session_dir = session["output_dir"]
    else:
        # Fallback to original paths if no session_id
        session_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'data')
    
    # Get the script's directory and build relative paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.join(script_dir, '..', '..')
    
    # Input files in session directory (look for Embankment Height file)
    emb_height_file = None
    for filename in os.listdir(session_dir):
        if 'emb_height' in filename.lower() or 'embankment' in filename.lower():
            emb_height_file = os.path.join(session_dir, filename)
            break
    
    if not emb_height_file:
        print("[ERROR] No Embankment Height file found in session directory")
        return
    
    # Main carriageway file in session directory with session_id suffix
    main_carriageway_file = os.path.join(session_dir, f'main_carriageway_{session_id}.xlsx')
    output_file = os.path.join(session_dir, f'main_carriageway_{session_id}.xlsx')
    
    try:
        # Step 1: Calculate embankment heights
        calculate_embankment_heights(emb_height_file, main_carriageway_file, output_file)
        
        print("\n" + "="*80)
        print("SUCCESS! [OK]")
        print("="*80)
        print("Output file:", output_file)
        print("Columns updated: BJ, BK, BL, BM (Embankment Heights)")
        
    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()

def wait_for_file_ready(file_path, max_wait_seconds=15):
    """Wait for file to be ready for access"""
    import os
    start_time = time.time()
    
    print(f"[INFO] Waiting for file to be ready: {os.path.basename(file_path)}")
    
    while time.time() - start_time < max_wait_seconds:
        try:
            # Try to open file in read mode to check if it's ready
            with open(file_path, 'rb') as f:
                pass
            # Additional check: try to read first few bytes
            with open(file_path, 'rb') as f:
                f.read(1024)  # Try to read first 1KB
            print(f"[OK] File is ready for access: {os.path.basename(file_path)}")
            return True
        except (IOError, PermissionError) as e:
            print(f"[WAIT] File not ready, waiting... ({time.time() - start_time:.1f}s elapsed)")
            time.sleep(1.0)
        except Exception as e:
            print(f"[ERROR] Unexpected error waiting for file: {e}")
            return False
    
    print(f"[ERROR] Timeout waiting for file to be ready: {file_path}")
    return False

def safe_workbook_operation(file_path, operation_func, max_retries=5):
    """Safely perform workbook operations with enhanced retry logic"""
    import os
    from openpyxl import load_workbook
    
    for attempt in range(max_retries):
        wb = None
        try:
            # Wait for file to be ready
            if not wait_for_file_ready(file_path, max_wait_seconds=10):
                raise IOError(f"File not ready for access: {file_path}")
            
            print(f"[INFO] Attempting workbook operation (attempt {attempt + 1}/{max_retries})")
            
            # Load workbook
            wb = load_workbook(file_path)
            result = operation_func(wb)
            
            # Save workbook
            wb.save(file_path)
            
            print(f"[OK] Workbook operation completed successfully")
            return result
            
        except Exception as e:
            print(f"[WARNING] Attempt {attempt + 1} failed: {e}")
            
            # Ensure workbook is closed even on error
            if wb:
                try:
                    wb.close()
                except:
                    pass
            
            if attempt == max_retries - 1:
                print(f"[ERROR] All attempts failed for {os.path.basename(file_path)}")
                raise e
            
            # Exponential backoff with jitter
            wait_time = (2 ** attempt) * 0.5 + (attempt * 0.1)
            print(f"[INFO] Waiting {wait_time:.1f}s before retry...")
            time.sleep(wait_time)
        
        finally:
            # Always close workbook
            if wb:
                try:
                    wb.close()
                except:
                    pass
            
            # Wait after operation completion
            time.sleep(1.0)

def calculate_embankment_heights(emb_height_file, main_carriageway_file, output_file):
    """
    Reads EMB_HEIGHT.xlsx and main_carriageway.xlsx
    Calculates embankment heights based on level differences
    Updates columns BJ, BK, BL, BM in main_carriageway.xlsx
    """
    print("="*80)
    print("EMBANKMENT HEIGHT CALCULATOR")
    print("="*80)
    
    # Read embankment height data
    print("\nSTEP 1: Reading Embankment Height Data")
    print("-" * 40)
    
    try:
        # Wait for embankment file to be ready
        if not wait_for_file_ready(emb_height_file, max_wait_seconds=10):
            raise IOError(f"Embankment height file not ready: {emb_height_file}")
        
        # Read columns B, C, D from EMB_HEIGHT.xlsx starting from row 7
        df_emb = pd.read_excel(emb_height_file, sheet_name='Emb Height', skiprows=6, usecols='B:D', header=None)
        df_emb.columns = ['Chainage', 'Existing_Level', 'Proposed_Level']
        
        # Remove empty rows
        df_emb = df_emb.dropna(how='all')
        
        # Convert to numeric
        df_emb['Chainage'] = pd.to_numeric(df_emb['Chainage'], errors='coerce')
        df_emb['Existing_Level'] = pd.to_numeric(df_emb['Existing_Level'], errors='coerce')
        df_emb['Proposed_Level'] = pd.to_numeric(df_emb['Proposed_Level'], errors='coerce')
        
        print(f"[OK] Read {len(df_emb)} rows from embankment height file")
        print(f"[OK] Chainage range: {df_emb['Chainage'].min():.3f} to {df_emb['Chainage'].max():.3f}")
        
    except Exception as e:
        print(f"[ERROR] Failed to read embankment height file: {e}")
        raise
    
    # Load and process main carriageway using safe operation
    print("\nSTEP 2: Processing Main Carriageway Data")
    print("-" * 40)
    
    def process_workbook(wb):
        """Inner function to process the workbook"""
        from openpyxl import load_workbook
        
        ws = wb['Quantity']
        
        print(f"[OK] Loaded workbook: {main_carriageway_file}")
        print(f"  Sheet: Quantity")
        print(f"  Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        # Define column indices (Excel columns, 1-indexed)
        FROM_COL = 1    # Column A
        TO_COL = 2      # Column B
        LENGTH_COL = 3   # Column C
        
        BJ_COL = 62     # Column BJ
        BK_COL = 63     # Column BK
        BL_COL = 64     # Column BL
        BM_COL = 65     # Column BM
        
        # Data starts from row 7
        start_row = 7
        
        # Find last row with data in column C
        last_row = 0
        for row_idx in range(start_row, ws.max_row + 1):
            cell_value = ws.cell(row_idx, LENGTH_COL).value
            if cell_value is not None and cell_value != '':
                last_row = row_idx
        
        if last_row == 0:
            print("[WARNING] No data found in column C")
            return 0
        
        print(f"[OK] Processing rows {start_row} to {last_row}")
        
        # Process each row in main carriageway
        processed_rows = 0
        for row_idx in range(start_row, last_row + 1):
            # Get chainage range
            from_chainage = ws.cell(row_idx, FROM_COL).value
            to_chainage = ws.cell(row_idx, TO_COL).value
            length = ws.cell(row_idx, LENGTH_COL).value
            
            # Skip empty rows
            if from_chainage is None or to_chainage is None or length is None or length == 0:
                continue
            
            try:
                from_chainage = float(from_chainage)
                to_chainage = float(to_chainage)
            except (ValueError, TypeError):
                continue
            
            # Calculate embankment heights at start and end of section
            start_height = get_embankment_height(df_emb, from_chainage)
            end_height = get_embankment_height(df_emb, to_chainage)
            
            # Calculate average height for section
            avg_height = (start_height + end_height) / 2 if start_height and end_height else 0
            
            # Calculate volume components
            volume_per_meter = avg_height if avg_height else 0
            total_volume = volume_per_meter * length if length else 0
            
            # Write to columns BJ, BK, BL, BM
            ws.cell(row_idx, BJ_COL).value = start_height if start_height else 0
            ws.cell(row_idx, BK_COL).value = end_height if end_height else 0
            ws.cell(row_idx, BL_COL).value = volume_per_meter if volume_per_meter else 0
            ws.cell(row_idx, BM_COL).value = total_volume if total_volume else 0
            
            processed_rows += 1
            
            # Progress indicator
            if processed_rows % 200 == 0:
                print(f"  Processed {processed_rows} rows...")
        
        print(f"[OK] Processed {processed_rows} rows")
        return processed_rows
    
    # Use safe workbook operation
    try:
        return safe_workbook_operation(main_carriageway_file, process_workbook)
    except Exception as e:
        print(f"[ERROR] Failed to process main carriageway workbook: {e}")
        raise

def get_embankment_height(df_emb, chainage):
    """
    Get embankment height at a specific chainage using linear interpolation
    """
    if df_emb.empty:
        return None
    
    # Find nearest embankment data points
    # Look for points before and after chainage
    before_point = None
    after_point = None
    
    for idx, row in df_emb.iterrows():
        if pd.isna(row['Chainage']):
            continue
            
        emb_chainage = float(row['Chainage'])
        
        if emb_chainage <= chainage:
            before_point = row
        if emb_chainage >= chainage and after_point is None:
            after_point = row
            break
    
    # Handle edge cases
    if before_point is None:
        if after_point is None:
            return None
        return float(after_point['Proposed_Level']) - float(after_point['Existing_Level'])
    
    if after_point is None:
        return float(before_point['Proposed_Level']) - float(before_point['Existing_Level'])
    
    # Linear interpolation
    if before_point['Chainage'] == after_point['Chainage']:
        height_diff = float(after_point['Proposed_Level']) - float(after_point['Existing_Level'])
    else:
        # Interpolate between points
        ratio = (chainage - float(before_point['Chainage'])) / (float(after_point['Chainage']) - float(before_point['Chainage']))
        
        start_height = float(before_point['Proposed_Level']) - float(before_point['Existing_Level'])
        end_height = float(after_point['Proposed_Level']) - float(after_point['Existing_Level'])
        
        height_diff = start_height + ratio * (end_height - start_height)
    
    return max(0, height_diff)  # Ensure non-negative height

if __name__ == "__main__":
    main()
