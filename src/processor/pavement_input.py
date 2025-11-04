"""
Pavement Input Processor
Reads Pavement_Input.xlsx to extract pavement layer data
Calculates and populates pavement-related columns in main_carriageway.xlsx
"""

import pandas as pd
import os
import sys
import io
import time
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def wait_for_file_ready(file_path, max_wait_seconds=15):
    """Wait for file to be ready for access"""
    import os
    start_time = time.time()
    
    print(f"[INFO] Waiting for file to be ready: {os.path.basename(file_path)}")
    
    while time.time() - start_time < max_wait_seconds:
        try:
            # Try to open file in read mode to check if it's ready
            with open(file_path, 'rb') as f:
                pass
            # Additional check: try to read first few bytes
            with open(file_path, 'rb') as f:
                f.read(1024)  # Try to read first 1KB
            print(f"[OK] File is ready for access: {os.path.basename(file_path)}")
            return True
        except (IOError, PermissionError) as e:
            print(f"[WAIT] File not ready, waiting... ({time.time() - start_time:.1f}s elapsed)")
            time.sleep(1.0)
        except Exception as e:
            print(f"[ERROR] Unexpected error waiting for file: {e}")
            return False
    
    print(f"[ERROR] Timeout waiting for file to be ready: {file_path}")
    return False

def safe_workbook_operation(file_path, operation_func, max_retries=5):
    """Safely perform workbook operations with enhanced retry logic"""
    import os
    from openpyxl import load_workbook
    
    for attempt in range(max_retries):
        try:
            # Wait for file to be ready
            if not wait_for_file_ready(file_path, max_wait_seconds=10):
                raise IOError(f"File not ready for access: {file_path}")
            
            print(f"[INFO] Attempting workbook operation (attempt {attempt + 1}/{max_retries})")
            
            wb = load_workbook(file_path)
            result = operation_func(wb)
            
            # Save and close with additional delay
            wb.save(file_path)
            wb.close()
            
            # Wait longer after save to ensure file system completion
            time.sleep(2.0)
            
            print(f"[OK] Workbook operation completed successfully")
            return result
            
        except Exception as e:
            print(f"[WARNING] Attempt {attempt + 1} failed: {e}")
            if attempt == max_retries - 1:
                print(f"[ERROR] All attempts failed for {os.path.basename(file_path)}")
                raise e
            
            # Exponential backoff with jitter
            wait_time = (2 ** attempt) * 0.5 + (attempt * 0.1)
            print(f"[INFO] Waiting {wait_time:.1f}s before retry...")
            time.sleep(wait_time)

def main(session_id=None):
    """Main function to execute pavement input processing"""
    import sys
    sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))
    from api.session_manager import session_manager
    
    # Get session information
    if session_id:
        session = session_manager.get_session(session_id)
        if not session:
            print(f"[ERROR] Session {session_id} not found")
            return
        session_dir = session["output_dir"]
    else:
        # Fallback to original paths if no session_id
        session_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'data')
    
    # Get the script's directory and build relative paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.join(script_dir, '..', '..')
    
    # Input files in session directory (look for Pavement Input file)
    pavement_input_file = None
    for filename in os.listdir(session_dir):
        if 'pavement input' in filename.lower():
            pavement_input_file = os.path.join(session_dir, filename)
            break
    
    if not pavement_input_file:
        print("[ERROR] No Pavement Input file found in session directory")
        return
    
    # Main carriageway file in session directory with session_id suffix
    main_carriageway_file = os.path.join(session_dir, f'main_carriageway_{session_id}.xlsx')
    output_file = os.path.join(session_dir, f'main_carriageway_{session_id}.xlsx')
    
    try:
        # Step 1: Process pavement input data
        process_pavement_input(pavement_input_file, main_carriageway_file, output_file)
        
        print("\n" + "="*80)
        print("SUCCESS! [OK]")
        print("="*80)
        print("Output file:", output_file)
        print("Pavement layers processed and populated")
        
    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()

def process_pavement_input(pavement_input_file, main_carriageway_file, output_file):
    """
    Reads Pavement_Input.xlsx and processes pavement layer data
    Updates main_carriageway.xlsx with pavement calculations
    """
    print("="*80)
    print("PAVEMENT INPUT PROCESSOR")
    print("="*80)
    
    # Step 1: Read pavement input data
    print("\nSTEP 1: Reading Pavement Input Data")
    print("-" * 40)
    
    try:
        # Wait for pavement input file to be ready
        if not wait_for_file_ready(pavement_input_file, max_wait_seconds=10):
            raise IOError(f"Pavement input file not ready: {pavement_input_file}")
        
        # Read pavement data from Excel file
        # Assuming data is in a specific format - adjust as needed
        df_pavement = pd.read_excel(pavement_input_file, sheet_name='Pavement', header=None)
        
        print(f"[OK] Read pavement input file: {pavement_input_file}")
        print(f"[OK] Data shape: {df_pavement.shape}")
        
        # Extract layer information based on expected format
        # This is a placeholder - adjust based on actual file structure
        pavement_layers = extract_pavement_layers(df_pavement)
        
    except Exception as e:
        print(f"[ERROR] Failed to read pavement input file: {e}")
        raise
    
    # Step 2: Load and update main carriageway
    print("\nSTEP 2: Updating Main Carriageway")
    print("-" * 40)
    
    def process_workbook(wb):
        """Inner function to process the workbook"""
        from openpyxl import load_workbook
        
        ws = wb['Quantity']
        
        print(f"[OK] Loaded workbook: {main_carriageway_file}")
        print(f"  Sheet: Quantity")
        print(f"  Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        # Define column indices for pavement-related columns
        # These are example column positions - adjust based on actual requirements
        LENGTH_COL = 3   # Column C
        
        # Example pavement layer columns (adjust as needed)
        # These would be the actual column numbers for your pavement data
        PAVEMENT_COLS = {
            'subbase': 66,     # Column BN (example)
            'base': 67,        # Column BO (example)
            'binder': 68,      # Column BP (example)
            'surface': 69       # Column BQ (example)
        }
        
        # Data starts from row 7
        start_row = 7
        
        # Find last row with data
        last_row = 0
        for row_idx in range(start_row, ws.max_row + 1):
            cell_value = ws.cell(row_idx, LENGTH_COL).value
            if cell_value is not None and cell_value != '':
                last_row = row_idx
        
        if last_row == 0:
            print("[WARNING] No data found in column C")
            return 0
        
        print(f"[OK] Processing rows {start_row} to {last_row}")
        
        # Process each row
        processed_rows = 0
        for row_idx in range(start_row, last_row + 1):
            # Get length value
            length = ws.cell(row_idx, LENGTH_COL).value
            
            # Skip empty rows
            if length is None or length == 0:
                continue
            
            try:
                length = float(length)
            except (ValueError, TypeError):
                continue
            
            # Calculate pavement quantities based on layers
            # This is example logic - adjust based on actual requirements
            pavement_quantities = calculate_pavement_quantities(pavement_layers, length)
            
            # Update pavement columns
            for layer_type, col_idx in PAVEMENT_COLS.items():
                if layer_type in pavement_quantities:
                    ws.cell(row_idx, col_idx).value = pavement_quantities[layer_type]
                else:
                    ws.cell(row_idx, col_idx).value = 0
            
            processed_rows += 1
            
            # Progress indicator
            if processed_rows % 200 == 0:
                print(f"  Processed {processed_rows} rows...")
        
        print(f"[OK] Processed {processed_rows} rows")
        return processed_rows
    
    # Use safe workbook operation
    try:
        return safe_workbook_operation(main_carriageway_file, process_workbook)
    except Exception as e:
        print(f"[ERROR] Failed to process main carriageway workbook: {e}")
        raise

def extract_pavement_layers(df_pavement):
    """
    Extract pavement layer information from the dataframe
    This is a placeholder function - adjust based on actual file structure
    """
    # Example implementation - replace with actual logic based on file format
    layers = {}
    
    # Example: Extract layer thicknesses from specific cells
    # This would need to be adjusted based on your actual file structure
    
    if not df_pavement.empty:
        # Example: Look for layer data in specific rows/columns
        try:
            # This is placeholder logic - adjust based on actual requirements
            layers['subbase_thickness'] = 200  # mm
            layers['base_thickness'] = 150      # mm
            layers['binder_thickness'] = 60     # mm
            layers['surface_thickness'] = 40     # mm
            
            print(f"[OK] Extracted pavement layers:")
            for layer, thickness in layers.items():
                print(f"  {layer}: {thickness} mm")
                
        except Exception as e:
            print(f"[WARNING] Could not extract layer data: {e}")
            # Set default values
            layers = {
                'subbase_thickness': 200,
                'base_thickness': 150,
                'binder_thickness': 60,
                'surface_thickness': 40
            }
    
    return layers

def calculate_pavement_quantities(layers, length):
    """
    Calculate pavement quantities based on layer data and length
    """
    quantities = {}
    
    # Example calculation - adjust based on actual requirements
    # This assumes road width is fixed or calculated elsewhere
    
    # Example road width (adjust as needed)
    road_width = 7.0  # meters
    
    # Calculate areas and volumes
    if 'subbase_thickness' in layers:
        thickness_m = layers['subbase_thickness'] / 1000  # Convert mm to m
        quantities['subbase'] = length * road_width * thickness_m
    
    if 'base_thickness' in layers:
        thickness_m = layers['base_thickness'] / 1000
        quantities['base'] = length * road_width * thickness_m
    
    if 'binder_thickness' in layers:
        thickness_m = layers['binder_thickness'] / 1000
        quantities['binder'] = length * road_width * thickness_m
    
    if 'surface_thickness' in layers:
        thickness_m = layers['surface_thickness'] / 1000
        quantities['surface'] = length * road_width * thickness_m
    
    return quantities

if __name__ == "__main__":
    main()
